from __future__ import annotations

import gc
import json
import os
import time
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from diagramador_optimizado.utils.time_utils import _to_minutes, formatear_hora


def generar_datos_de_simulacion(simular_excel: bool = True) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """
    Genera una configuración y viajes sintéticos para pruebas rápidas.

    Args:
        simular_excel: Si es True, devuelve una lista de viajes comerciales ejemplo.

    Returns:
        Tupla (config, viajes). Los viajes son una lista vacía cuando simular_excel es False.
    """
    print("FASE 0: Cargando datos de simulación (JSON y Excel)...")
    config = {
        "deposito": "Deposito",
        "limite_jornada": 600,
        "tiempo_toma": 15,
        "costo_por_bus": 1000,
        "paradas": {
            "TILVOL": {"min": 5, "max": 20},
            "ROTPIR": {"min": 10, "max": 45},
        },
        "vacios": {
            "Deposito_TILVOL": {"franjas": [{"inicio": "00:00", "fin": "40:00", "tiempo": 30, "km": 5}]},
            "Deposito_ROTPIR": {"franjas": [{"inicio": "00:00", "fin": "07:00", "tiempo": 65, "km": 10}, {"inicio": "07:01", "fin": "40:00", "tiempo": 60, "km": 10}]},
            "TILVOL_Deposito": {"franjas": [{"inicio": "00:00", "fin": "40:00", "tiempo": 30, "km": 5}]},
            "ROTPIR_Deposito": {"franjas": [{"inicio": "00:00", "fin": "40:00", "tiempo": 60, "km": 10}]},
            "TILVOL_ROTPIR": {"franjas": [{"inicio": "00:00", "fin": "40:00", "tiempo": 60, "km": 12}]},
        },
        "desplazamientos": {
            "Deposito_TILVOL": {"habilitado": True, "tiempo": 30},
            "TILVOL_Deposito": {"habilitado": True, "tiempo": 30},
            "Deposito_ROTPIR": {"habilitado": False, "tiempo": 60},
            "ROTPIR_Deposito": {"habilitado": False, "tiempo": 60},
            "TILVOL_ROTPIR": {"habilitado": True, "tiempo": 45},
            "ROTPIR_TILVOL": {"habilitado": True, "tiempo": 45},
        },
    }

    viajes = []
    if simular_excel:
        identificador = 0
        for minuto in range(6 * 60, 23 * 60 + 1, 10):
            viajes.append(
                {
                    "id": identificador,
                    "linea": "L1",
                    "sentido": "IDA",
                    "origen": "TILVOL",
                    "destino": "ROTPIR",
                    "inicio": minuto,
                    "fin": minuto + 60,
                    "kilometros": 12.0,
                    "desc": f"L1 IDA {formatear_hora(minuto)} TILVOL->ROTPIR",
                }
            )
            identificador += 1
            viajes.append(
                {
                    "id": identificador,
                    "linea": "L1",
                    "sentido": "VUELTA",
                    "origen": "ROTPIR",
                    "destino": "TILVOL",
                    "inicio": minuto,
                    "fin": minuto + 60,
                    "kilometros": 12.0,
                    "desc": f"L1 VUELTA {formatear_hora(minuto)} ROTPIR->TILVOL",
                }
            )
            identificador += 1
        print(f"Datos simulados: {len(viajes)} viajes a cubrir.")

    return config, viajes


def cargar_config(ruta_json: str = "configuracion.json") -> Dict[str, Any]:
    """
    Carga el archivo de configuración en disco o recurre a datos simulados.

    Args:
        ruta_json: Ruta absoluta o relativa al archivo JSON.

    Returns:
        Diccionario con la configuración normalizada.
    """
    time.sleep(0.3)
    gc.collect()

    try:
        contenido = None
        for intento in range(3):
            try:
                with open(ruta_json, "r", encoding="utf-8") as archivo:
                    contenido = archivo.read()
                if contenido and contenido.strip():
                    break
                print(f"Intento {intento + 1}: Archivo vacío, reintentando...")
                time.sleep(0.1)
            except Exception as error_intento:
                print(f"Intento {intento + 1}: Error leyendo archivo: {error_intento}")
                if intento < 2:
                    time.sleep(0.1)
                else:
                    raise

        if not contenido or not contenido.strip():
            raise ValueError(f"El archivo {ruta_json} está vacío o no se pudo leer.")

        config = json.loads(contenido)
        if not isinstance(config, dict):
            raise ValueError(f"El archivo {ruta_json} no contiene un diccionario válido.")

        print("=" * 80)
        print(f"CONFIGURACIÓN CARGADA DESDE DISCO: {ruta_json}")
        print(f"  - Depósito: {config.get('deposito', 'NO ENCONTRADO')}")
        print(f"  - Nodos: {config.get('nodos', [])} (total: {len(config.get('nodos', []))})")
        print(f"  - Valor RAW de 'limite_jornada': {config.get('limite_jornada', 'NO ENCONTRADO')}")
        print(f"=" * 80)

    except FileNotFoundError:
        raise FileNotFoundError(f"No se encontró el archivo de configuración: {ruta_json}")
    except json.JSONDecodeError as error_json:
        raise ValueError(f"JSON inválido en {ruta_json}: {error_json}") from error_json
    except Exception as error_general:
        raise RuntimeError(f"Error al cargar configuración desde {ruta_json}: {error_general}") from error_general

    # Normalizar claves esenciales.
    config.setdefault("deposito", "Deposito")
    config.setdefault("limite_jornada", 720)
    config.setdefault("tiempo_toma", 15)
    config.setdefault("max_buses", 200)
    config.setdefault("paradas", {})
    config.setdefault("vacios", {})
    config.setdefault("desplazamientos", {})
    config.setdefault("nodos", [])
    config.setdefault("tipos_bus", {})
    config.setdefault("lineas", {})
    config.setdefault("flota_por_tipo", {})
    config.setdefault("max_cambios_bus_conductor", 2)
    config.setdefault("tipos_conductor", [])

    return config


def cargar_salidas_desde_excel(ruta_excel: str) -> List[Dict[str, Any]]:
    """
    Lee un archivo Excel y devuelve la lista de viajes comerciales normalizados.

    Args:
        ruta_excel: Ruta al archivo XLSX.

    Returns:
        Lista de diccionarios representando viajes comerciales.
    """
    try:
        libro = load_workbook(ruta_excel, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(f"No se encontró Excel en {ruta_excel}")

    hoja = libro.active
    try:
        encabezados = [celda.value for celda in hoja[1]]
        indices = {nombre: idx for idx, nombre in enumerate(encabezados)}
        requeridos = ["Linea", "Sentido", "Origen", "Destino", "Hora Inicio", "Hora Fin", "Kilometros"]
        for campo in requeridos:
            if campo not in indices:
                raise ValueError(f"Falta columna obligatoria en Excel: {campo}")
    except Exception as error_headers:
        raise ValueError(f"Error leyendo encabezados del Excel: {error_headers}") from error_headers

    viajes: List[Dict[str, Any]] = []
    identificador = 0
    vistos_por_clave: Dict[Tuple[str, str, str, str, int, int], int] = {}

    for fila_idx, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
        if not fila or fila[indices["Origen"]] is None or fila[indices["Destino"]] is None:
            continue
        try:
            linea = str(fila[indices["Linea"]] or "")
            sentido = str(fila[indices["Sentido"]] or "")
            origen = str(fila[indices["Origen"]])
            destino = str(fila[indices["Destino"]])
            inicio = _to_minutes(fila[indices["Hora Inicio"]])
            fin = _to_minutes(fila[indices["Hora Fin"]])
            if inicio is None or fin is None:
                raise ValueError("Hora Inicio/Hora Fin inválida o vacía.")
            if fin < inicio:
                raise ValueError(
                    f"Hora Fin ({formatear_hora(fin)}) < Hora Inicio ({formatear_hora(inicio)}). "
                    "No se permiten autocorrecciones."
                )
            try:
                kilometros = float(str(fila[indices["Kilometros"]]).replace(",", ".")) if fila[indices["Kilometros"]] is not None else 0.0
            except Exception as error_km:
                raise ValueError(f"Kilometros inválido: {fila[indices['Kilometros']]}") from error_km
            descripcion = f"{linea} {sentido} {formatear_hora(inicio)} {origen}->{destino}"

            clave_viaje = (
                linea.strip().upper(),
                sentido.strip().upper(),
                origen.strip().upper(),
                destino.strip().upper(),
                int(inicio),
                int(fin),
            )
            fila_prev = vistos_por_clave.get(clave_viaje)
            if fila_prev is not None:
                raise ValueError(
                    f"Viaje duplicado detectado. Fila actual {fila_idx}, fila previa {fila_prev}. "
                    f"Clave={clave_viaje}"
                )
            vistos_por_clave[clave_viaje] = fila_idx

            viajes.append(
                {
                    "id": identificador,
                    "linea": linea,
                    "sentido": sentido,
                    "origen": origen,
                    "destino": destino,
                    "inicio": inicio,
                    "fin": fin,
                    "kilometros": kilometros,
                    "desc": descripcion,
                }
            )
            identificador += 1
        except Exception as error_fila:
            raise ValueError(f"Error procesando fila {fila_idx}: {error_fila}") from error_fila

    return viajes

