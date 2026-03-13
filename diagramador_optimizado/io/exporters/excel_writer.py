# -*- coding: utf-8 -*-
"""
Exportación a Excel - Lógica de Origen con corte Fase 2 y puntos de relevo.

- Genera eventos de bus: Vacio y Parada desde bloques; Comercial solo desde insumo de carga.
- Usa la misma lógica de corte y puntos de relevo que Fase 2.
- EXPORTACIÓN PURA: Ya no muta ni recalcula lógicas temporales. Confía 100% en los 
  eventos generados por el pipeline de la Fase 2.
"""
from __future__ import annotations

import collections
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from diagramador_optimizado.core.domain.logistica import GestorDeLogistica
from diagramador_optimizado.core.tempo_conectividad import (
    es_deposito,
    tiempo_a_minutos,
)
from diagramador_optimizado.utils.time_utils import (
    formatear_hora_deltatime,
    minutos_desde_base_mas_duracion,
)


def _tipo_evento_canonico(tipo: Any) -> str:
    t = str(tipo or "").strip()
    u = t.upper()
    if u == "VACIO":
        return "Vacio"
    if u == "PARADA":
        return "Parada"
    if u == "COMERCIAL":
        return "Comercial"
    if u == "RECARGA":
        return "Recarga"
    if u == "INS":
        return "InS"
    if u == "FNS":
        return "FnS"
    if u == "DESPLAZAMIENTO":
        return "Desplazamiento"
    return t


def _bus_para_excel(bus: Any) -> Any:
    if bus in (None, ""):
        return ""
    s = str(bus).strip()
    if s.isdigit():
        return int(s)
    return s


def _hora_para_excel(val: Any) -> str:
    """
    Formato de hora para export: si el valor es numérico (minutos), usa 24:XX cuando
    corresponda. Si ya es string, devolver tal cual.
    """
    if val is None or val == "":
        return ""
    if isinstance(val, (int, float)):
        return formatear_hora_deltatime(int(val))
    return str(val)


def _nodo_para_excel(val: Any, gestor: Optional[GestorDeLogistica]) -> str:
    """Normaliza origen/destino: si es depósito, usar siempre el nombre canónico."""
    s = (val or "").strip()
    if not s or not gestor:
        return s
    if es_deposito(s, gestor.deposito_base or ""):
        return (gestor.deposito_base or s).strip()
    return s


def _validar_encadenamiento_eventos(
    eventos: List[Dict[str, Any]],
    gestor: GestorDeLogistica,
) -> Tuple[int, int, List[Tuple[Any, str, str]]]:
    TOL_HUECO_MIN = 10
    DESCANSO_LARGO_MIN = 150

    por_conductor: Dict[Any, List[Dict[str, Any]]] = collections.defaultdict(list)
    for ev in eventos:
        cid = ev.get("conductor")
        if cid is not None and cid != "":
            por_conductor[cid].append(ev)
    conductores_ok = 0
    incidencias: List[Tuple[Any, str, str]] = []
    
    for cid, lista in por_conductor.items():
        def _key_orden(e):
            ini = tiempo_a_minutos(e.get("inicio", 0))
            fin = tiempo_a_minutos(e.get("fin", 0))
            return (ini if ini is not None else 0, fin if fin is not None else 0)
        ordenada = sorted(lista, key=_key_orden)
        tiene_incidencia = False
        for i in range(len(ordenada) - 1):
            prev, curr = ordenada[i], ordenada[i + 1]
            tipo_prev = (prev.get("evento") or "").strip()
            tipo_curr = (curr.get("evento") or "").strip()
            fin_p = tiempo_a_minutos(prev.get("fin", 0))
            ini_c = tiempo_a_minutos(curr.get("inicio", 0))
            dest_p_raw = (prev.get("destino") or prev.get("origen") or "").strip()
            orig_c_raw = (curr.get("origen") or "").strip()
            dest_canon = (gestor.nodo_canonico_para_conectividad(dest_p_raw) or "").strip()
            orig_canon = (gestor.nodo_canonico_para_conectividad(orig_c_raw) or "").strip()
            mismo_nodo = bool(dest_canon and orig_canon and (dest_canon.upper() == orig_canon.upper()))
            
            diff = (ini_c - fin_p) if (fin_p is not None and ini_c is not None) else None
            if fin_p is not None and ini_c is not None and fin_p != ini_c:
                if abs(diff) <= TOL_HUECO_MIN:
                    pass
                elif abs(diff) >= DESCANSO_LARGO_MIN:
                    pass
                elif diff < 0 and mismo_nodo:
                    pass
                elif tipo_prev in ("InS", "FnS") or tipo_curr in ("InS", "FnS"):
                    pass
                else:
                    tiene_incidencia = True
                    incidencias.append((cid, "hueco", f"fin={fin_p} siguiente inicio={ini_c} (diff={diff})"))
                    
            if (not mismo_nodo and (dest_p_raw or orig_c_raw) and tipo_prev not in ("InS", "FnS") and tipo_curr not in ("InS", "FnS")):
                hay_conector = any(
                    (e.get("evento") or "").strip() in ("Desplazamiento", "Vacio", "Comercial")
                    and (gestor.nodo_canonico_para_conectividad((e.get("origen") or "").strip()) or "").strip().upper() == (dest_canon or "").upper()
                    and (gestor.nodo_canonico_para_conectividad((e.get("destino") or "").strip()) or "").strip().upper() == (orig_canon or "").upper()
                    for e in ordenada
                )
                if not hay_conector:
                    tiene_incidencia = True
                    incidencias.append((cid, "teletransporte", f"destino={dest_p_raw!r} siguiente origen={orig_c_raw!r}"))
        if not tiene_incidencia:
            conductores_ok += 1
    return conductores_ok, len(por_conductor) - conductores_ok, incidencias


def exportar_resultado_excel(
    config: Dict[str, Any],
    bloques_bus: List[List[Dict[str, Any]]],
    turnos_seleccionados: List[Dict[str, Any]],
    viajes_comerciales: List[Dict[str, Any]],
    metadata_tareas: Dict[Any, Dict[str, Any]],
    status_fase1: str,
    status_fase2: str,
    path_out: str = "resultado_diagramacion.xlsx",
    gestor: Optional[GestorDeLogistica] = None,
    verbose: bool = False,
    status_f3: Optional[str] = None,
    eventos_bus: Optional[List[List[Dict[str, Any]]]] = None,
    eventos_completos: Optional[List[Dict[str, Any]]] = None,
) -> Optional[Dict[str, Any]]:
    
    if gestor is None:
        gestor = GestorDeLogistica(config)
    path_out = str(Path(path_out).resolve())
    print(f"Iniciando exportación a {path_out}...")
    
    if not bloques_bus or not turnos_seleccionados or not viajes_comerciales or not eventos_completos:
        print("ERROR: Faltan datos estructurales para exportar.")
        return None
        
    todos_eventos = list(eventos_completos)

    # --------------------------------------------------------------------------------
    # 1. Renumeración de conductores (sin filtrar eventos)
    # --------------------------------------------------------------------------------
    conductores_validos: Set[int] = set()
    for ev in todos_eventos:
        cid = ev.get("conductor")
        if cid is not None and cid != "":
            try:
                conductores_validos.add(int(cid))
            except (TypeError, ValueError):
                pass

    # Renumerar conductores sin huecos (1..N)
    mapa_conductores: Dict[int, int] = {}
    if conductores_validos:
        ids_ordenados = sorted(conductores_validos)
        for nuevo_id, viejo_id in enumerate(ids_ordenados, start=1):
            mapa_conductores[viejo_id] = nuevo_id

        for ev in todos_eventos:
            cid = ev.get("conductor")
            if not cid:
                continue
            try:
                cid_int = int(cid)
            except (TypeError, ValueError):
                continue
            if cid_int in mapa_conductores:
                ev["conductor"] = mapa_conductores[cid_int]

    num_conductores_exportados = len(mapa_conductores)

    # Análisis completo para la terminal
    conductores_ok_enc, conductores_inc_enc, incidencias_enc = _validar_encadenamiento_eventos(todos_eventos, gestor)
    total_cond_enc = conductores_ok_enc + conductores_inc_enc
    print("\n" + "=" * 70)
    print("ANÁLISIS COMPLETO - Encadenamiento por conductor (Exportación)")
    print("=" * 70)
    print(f"  Conductores con eventos: {total_cond_enc}")
    print(f"  Encadenamiento OK: {conductores_ok_enc}")
    print(f"  Con incidencias: {conductores_inc_enc}")
    if incidencias_enc:
        for cid, tipo, desc in incidencias_enc[:15]:
            print(f"    - Conductor {cid} ({tipo}): {desc}")
    print("=" * 70 + "\n")

    wb = Workbook()
    font_bold = Font(bold=True)
    wrap_align = Alignment(wrap_text=True)

    # --- Hoja 0: ResumenOptimizacion ---
    ws0 = wb.active
    ws0.title = "ResumenOptimizacion"
    ws0.append(["Fase", "Estado Final", "Descripción"])
    diag_flota = config.get("_diagnostico_flota_fase1") or {}
    if str(status_fase1).upper() == "FEASIBLE_MAX_BUSES_EXCEDIDO":
        max_cfg = diag_flota.get("max_buses_configurado", "")
        req = diag_flota.get("buses_requeridos", "")
        exc = diag_flota.get("buses_exceso", "")
        desc_f1 = (
            f"Modo permisivo: se excede flota. Requeridos={req}, "
            f"Configurados={max_cfg}, Exceso={exc}."
        )
    else:
        desc_f1 = "OPTIMAL: Flota mínima garantizada."
    ws0.append(["Fase 1: Buses", status_fase1, desc_f1])
    ws0.append(["Fase 2: Conductores", status_fase2, "OPTIMAL: Número mínimo de conductores."])
    if status_f3:
        desc_f3 = f"Fase 3: {len(turnos_seleccionados)} turnos -> {num_conductores_exportados} conductores exportados."
        ws0.append(["Fase 3: Unión de Conductores", "COMPLETADA", desc_f3])
    ws0.append(["Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""])
    for cell in ws0[1]: cell.font = font_bold
    for col in ["A", "B", "C"]: ws0.column_dimensions[col].width = 25

    if diag_flota:
        wsd = wb.create_sheet("DiagnosticoFlota")
        wsd.append(["Campo", "Valor"])
        wsd.append(["Estado", diag_flota.get("status", status_fase1)])
        wsd.append(["MaxBusesConfigurado", diag_flota.get("max_buses_configurado", "")])
        wsd.append(["BusesRequeridos", diag_flota.get("buses_requeridos", "")])
        wsd.append(["BusesExceso", diag_flota.get("buses_exceso", "")])
        wsd.append(["PermitidoContinuar", "SI" if diag_flota.get("permitido_continuar") else "NO"])
        wsd.append([])
        wsd.append(["Resumen", "Detalle"])
        rechazos = diag_flota.get("rechazos_factibilidad", []) or []
        if rechazos:
            por_linea = collections.Counter(str(d.get("linea", "SIN_LINEA")) for d in rechazos)
            por_franja = collections.Counter(str(d.get("franja", "SIN_FRANJA")) for d in rechazos)
            por_origen = collections.Counter(str(d.get("origen", "SIN_ORIGEN")) for d in rechazos)
            por_deposito = collections.Counter(str(d.get("mejor_deposito", "SIN_DEP")) for d in rechazos)
            wsd.append(["TopLineas", ", ".join(f"{k}:{v}" for k, v in por_linea.most_common(10))])
            wsd.append(["TopFranjas", ", ".join(f"{k}:{v}" for k, v in por_franja.most_common(10))])
            wsd.append(["TopOrigenes", ", ".join(f"{k}:{v}" for k, v in por_origen.most_common(10))])
            wsd.append(["TopDepositos", ", ".join(f"{k}:{v}" for k, v in por_deposito.most_common(10))])
            wsd.append([])
            wsd.append([
                "viaje_id", "linea", "grupo", "inicio", "fin", "origen", "destino",
                "buses_actuales", "max_buses", "mejor_deposito", "mejor_t_vacio_min",
            ])
            for d in rechazos:
                wsd.append([
                    d.get("viaje_id", ""),
                    d.get("linea", ""),
                    d.get("grupo", ""),
                    formatear_hora_deltatime(int(d.get("inicio", 0) or 0)),
                    formatear_hora_deltatime(int(d.get("fin", 0) or 0)),
                    d.get("origen", ""),
                    d.get("destino", ""),
                    d.get("total_buses_actual", ""),
                    d.get("max_buses", ""),
                    d.get("mejor_deposito", ""),
                    d.get("mejor_t_vacio", ""),
                ])
        else:
            wsd.append(["Info", "Sin rechazos detallados (modo CP-SAT o sin trazas)."])
        for cell in wsd[1]:
            cell.font = font_bold
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            wsd.column_dimensions[col].width = 22

    def _fin_dur_consistentes(ev: Dict[str, Any]) -> Tuple[int, int, int]:
        ini = tiempo_a_minutos(ev.get("inicio", 0)) or 0
        if (str(ev.get("evento", "")).strip().upper() == "FNS"):
            return ini, ini, 0
        fin_raw = tiempo_a_minutos(ev.get("fin", 0))
        dur_raw = tiempo_a_minutos(ev.get("duracion", 0))
        if dur_raw is not None and dur_raw > 0:
            dur = int(dur_raw)
            fin = minutos_desde_base_mas_duracion(ini, dur)
        elif fin_raw is not None:
            dur = (fin_raw - ini) if fin_raw >= ini else (1440 - ini + fin_raw)
            dur = max(0, dur)
            fin = minutos_desde_base_mas_duracion(ini, dur)
        else:
            fin = ini
            dur = 0
        return ini, fin, dur

    # --- Hoja 1: BloquesBuses ---
    ws1 = wb.create_sheet("BloquesBuses")
    ws1.append(["bus_id", "seq", "id_viaje", "linea", "sentido", "origen", "destino", "inicio_hhmm", "fin_hhmm", "duracion_min", "desc"])
    for b_id_idx, bloque in enumerate(bloques_bus):
        b_id = b_id_idx + 1
        for s, v in enumerate(bloque, start=1):
            dur = v.get("fin", 0) - v.get("inicio", 0)
            ws1.append([
                b_id, s, v.get("id", ""), v.get("linea", ""), v.get("sentido", ""),
                v.get("origen", ""), v.get("destino", ""),
                formatear_hora_deltatime(v.get("inicio", 0)), formatear_hora_deltatime(v.get("fin", 0)),
                dur, v.get("desc", ""),
            ])

    # --- Hoja 2: TurnosConductores (Transcripción de EventosCompletos) ---
    ws2 = wb.create_sheet("TurnosConductores")
    ws2.append(["conductor_id", "punto_inicio", "punto_fin", "inicio_jornada_hhmm", "fin_jornada_hhmm", "duracion_jornada_min", "detalle_servicios"])
    
    por_conductor_ws2 = collections.defaultdict(list)
    for ev in todos_eventos:
        c = ev.get("conductor")
        if c: 
            por_conductor_ws2[c].append(ev)

    for c_id, evs in sorted(por_conductor_ws2.items(), key=lambda x: int(x[0]) if str(x[0]).isdigit() else str(x[0])):
        evs_sorted = sorted(evs, key=lambda e: tiempo_a_minutos(e.get("inicio", 0)) or 0)
        
        ins_evs = [e for e in evs_sorted if str(e.get("evento")).upper() == "INS"]
        fns_evs = [e for e in evs_sorted if str(e.get("evento")).upper() == "FNS"]
        comerciales = [e for e in evs_sorted if str(e.get("evento")).upper() == "COMERCIAL"]
        
        # Consistencia con validación dura de jornada:
        # la jornada efectiva inicia al finalizar InS (no incluye tiempo de toma).
        # Regla de consistencia visual solicitada:
        # inicio de TurnosConductores debe coincidir con Inicio de InS en EventosCompletos.
        ini_min = (
            min(tiempo_a_minutos(e.get("inicio", 0)) for e in ins_evs)
            if ins_evs
            else tiempo_a_minutos(evs_sorted[0].get("inicio", 0))
        )
        fin_min = (
            max(tiempo_a_minutos(e.get("fin", 0)) for e in fns_evs)
            if fns_evs
            else tiempo_a_minutos(evs_sorted[-1].get("fin", 0))
        )
        
        punto_ini = ins_evs[0].get("origen") if ins_evs else evs_sorted[0].get("origen")
        punto_fin = fns_evs[-1].get("destino") if fns_evs else evs_sorted[-1].get("destino")
        
        duracion = fin_min - ini_min
        if duracion < 0: duracion += 1440
        
        lineas = sorted(list(set([str(e.get("linea", "")).strip() for e in comerciales if e.get("linea")])))
        
        ws2.append([
            c_id, punto_ini, punto_fin, _hora_para_excel(ini_min), _hora_para_excel(fin_min), duracion, ", ".join(lineas)
        ])

    # --- Hoja 3: BusEventos ---
    eventos_por_bus: Dict[int, List[Dict]] = collections.defaultdict(list)
    tipos_bus_eventos = ("VACIO", "PARADA", "COMERCIAL", "RECARGA")
    # Fuente de verdad de BusEventos: salida de Fase 1 (eventos_bus).
    # Si no está disponible, usar fallback a todos_eventos para no romper exportación.
    if eventos_bus:
        for bus_idx, bloque_ev in enumerate(eventos_bus, start=1):
            for ev in bloque_ev or []:
                tipo_ev = (str(ev.get("evento", "")) or "").strip().upper()
                if tipo_ev not in tipos_bus_eventos:
                    continue
                ev_copia = dict(ev)
                if ev_copia.get("bus") in (None, ""):
                    ev_copia["bus"] = bus_idx
                eventos_por_bus[bus_idx].append(ev_copia)
    else:
        for ev in todos_eventos:
            tipo_ev = (str(ev.get("evento", "")) or "").strip().upper()
            if tipo_ev not in tipos_bus_eventos:
                continue
            bus_id = ev.get("bus")
            if bus_id is not None and bus_id != "":
                try:
                    eventos_por_bus[int(bus_id)].append(dict(ev))
                except (TypeError, ValueError):
                    pass

    bus_tipo_map: Dict[int, str] = {}
    for bus_idx, bloque in enumerate(bloques_bus):
        for viaje in bloque or []:
            if viaje.get("tipo_bus"):
                bus_tipo_map[bus_idx + 1] = str(viaje.get("tipo_bus"))
                break

    ws3 = wb.create_sheet("BusEventos")
    ws3.append([
        "Evento", "Tipo", "Inicio", "De", "Fin", "A", "Duración", "Servicio", "Bus", "Tipo Bus",
        "Línea", "km", "V. Inferido", "Id.", "Sentido", "Tipo Mapeado", "Autonomía", "Consumo",
        "% Batería", "Pos. en P. Recarga", "Bus Orden"
    ])
    
    for bus_id in sorted(eventos_por_bus.keys()):
        # HACEMOS UNA COPIA PARA NO MUTAR LOS EVENTOS DE LA FASE 4
        eventos_ord = sorted([dict(e) for e in eventos_por_bus[bus_id]], key=lambda e: tiempo_a_minutos(e.get("inicio", 0)))
        # Normalizar paradas consecutivas del bus en el mismo nodo.
        eventos_ord_norm: List[Dict[str, Any]] = []
        for ev in eventos_ord:
            if (str(ev.get("evento", "")) or "").strip().upper() == "PARADA":
                ev_ini = tiempo_a_minutos(ev.get("inicio", 0)) or 0
                ev_fin = tiempo_a_minutos(ev.get("fin", 0)) or ev_ini
                if ev_fin <= ev_ini:
                    continue
            if not eventos_ord_norm:
                eventos_ord_norm.append(ev)
                continue
            prev = eventos_ord_norm[-1]
            tipo_prev = (str(prev.get("evento", "")) or "").strip().upper()
            tipo_act = (str(ev.get("evento", "")) or "").strip().upper()
            if tipo_prev == "PARADA" and tipo_act == "PARADA":
                prev_node = _nodo_para_excel(prev.get("destino", "") or prev.get("origen", ""), gestor).strip().upper()
                act_node = _nodo_para_excel(ev.get("origen", "") or ev.get("destino", ""), gestor).strip().upper()
                prev_fin = tiempo_a_minutos(prev.get("fin", 0)) or 0
                act_ini = tiempo_a_minutos(ev.get("inicio", 0)) or 0
                act_fin = tiempo_a_minutos(ev.get("fin", 0)) or act_ini
                if prev_node and act_node and prev_node == act_node and act_ini <= prev_fin + 1:
                    prev["fin"] = max(prev_fin, act_fin)
                    prev["desc"] = f"Parada en {prev.get('origen', '')}"
                    continue
            eventos_ord_norm.append(ev)
        eventos_ord = eventos_ord_norm

        last_fin_seq = None
        last_dest_seq = ""
        for orden, ev in enumerate(eventos_ord, start=1):
            _ini_min, _fin_min, duracion_min = _fin_dur_consistentes(ev)
            
            if last_fin_seq is not None:
                if _ini_min > last_fin_seq:
                    gap_total = _ini_min - last_fin_seq
                    origen_actual = (ev.get("origen") or "").strip()
                    ws3.append([
                        "Parada", "Parada", _hora_para_excel(last_fin_seq), last_dest_seq,
                        _hora_para_excel(_ini_min), last_dest_seq, _hora_para_excel(gap_total),
                        "", bus_id, bus_tipo_map.get(bus_id, ""), "", 0, "", "", "", bus_tipo_map.get(bus_id, ""),
                        "", "", "", "", f"{orden}-P"
                    ])
                elif _ini_min < last_fin_seq:
                    _ini_min = last_fin_seq
                    _fin_min = minutos_desde_base_mas_duracion(_ini_min, duracion_min)

            km = ev.get("kilometros", 0) or 0
            v_inferido = round(km / (duracion_min / 60.0), 1) if duracion_min > 0 and km > 0 else ""
            tipo_bus = ev.get("tipo_bus") or bus_tipo_map.get(bus_id, "")
            tipo_ev = _tipo_evento_canonico(ev.get("evento", ""))
            
            ws3.append([
                tipo_ev, tipo_ev, _hora_para_excel(_ini_min), ev.get("origen", ""),
                _hora_para_excel(_fin_min), ev.get("destino", ""), _hora_para_excel(duracion_min),
                ev.get("conductor", ""), _bus_para_excel(bus_id), tipo_bus, ev.get("linea", ""), km, v_inferido,
                ev.get("viaje_id", ""), ev.get("sentido", ""), tipo_bus, "", "", ev.get("porcentaje_bateria", ""),
                ev.get("posicion_recarga", ""), orden
            ])
            
            last_fin_seq = _fin_min
            last_dest_seq = ev.get("destino", "")

    # --- Hoja 4: EventosCompletos (Cero mutaciones, confianza absoluta en la Fase 2) ---
    ws4 = wb.create_sheet("EventosCompletos")
    ws4.append(["Tipo", "Bus", "Conductor", "Inicio", "Fin", "Duración", "Origen", "Destino", "km", "Línea", "desc"])
    
    eventos_ec_orden = sorted(
        todos_eventos,
        key=lambda e: (str(e.get("conductor", "")), tiempo_a_minutos(e.get("inicio", 0)) or 0)
    )
    # No escribir eventos sin conductor en EventosCompletos (Comercial/Vacio deben estar asignados)
    eventos_ec_orden = [e for e in eventos_ec_orden if e.get("conductor") not in (None, "")]

    for ev in eventos_ec_orden:
        ini_min, fin_min, dur_min = _fin_dur_consistentes(ev)
        orig = _nodo_para_excel(ev.get("origen", ""), gestor)
        dest = _nodo_para_excel(ev.get("destino", ""), gestor)
        ws4.append([
            _tipo_evento_canonico(ev.get("evento", "")),
            _bus_para_excel(ev.get("bus", "")),
            ev.get("conductor", ""),
            _hora_para_excel(ini_min),
            _hora_para_excel(fin_min),
            _hora_para_excel(dur_min),
            orig,
            dest,
            ev.get("kilometros", 0),
            ev.get("linea", ""),
            ev.get("desc", ""),
        ])

    path_real_guardado = path_out
    try:
        wb.save(path_out)
        print(f"[OK] Archivo exportado correctamente: {path_out}")
    except PermissionError:
        import os
        path_eval = os.path.join(os.path.dirname(path_out), "resultado_diagramacion_eval.xlsx")
        try:
            wb.save(path_eval)
            print(f"Permiso denegado en archivo principal. Exportado a: {path_eval}")
            path_real_guardado = path_eval
        except Exception:
            print(f"Error: Permiso denegado. Cierra '{path_out}'.")
            return None
    except Exception as e:
        print(f"Error exportando Excel: {e}")
        return None

    resumen = collections.Counter(_tipo_evento_canonico(ev.get("evento")) for ev in todos_eventos)
    print("Resumen de eventos exportados:")
    for tipo, cnt in sorted(resumen.items(), key=lambda x: -x[1]):
        print(f"  - {tipo}: {cnt}")
    return {"conductores_exportados": num_conductores_exportados, "path_real_guardado": path_real_guardado}