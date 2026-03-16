from __future__ import annotations

import os
import sys
from typing import Optional, Any, Dict, List, Tuple

from diagramador_optimizado.io.loaders import cargar_config, cargar_salidas_desde_excel
from diagramador_optimizado.io.config_validator import (
    validar_configuracion,
    ConfigValidationError,
    autocompletar_configuracion,
)
from diagramador_optimizado.io.exporters.excel_writer import exportar_resultado_excel
from diagramador_optimizado.core.domain.logistica import GestorDeLogistica
from diagramador_optimizado.core.engines.fase1_buses import resolver_diagramacion_buses
from diagramador_optimizado.core.engines.fase2_conductores import (
    resolver_diagramacion_conductores,
)
from diagramador_optimizado.core.engines.fase3_union import resolver_union_conductores
from diagramador_optimizado.core.engines.eventos_completos import construir_eventos_completos
from diagramador_optimizado.core.validaciones_fase import (
    validar_turnos_limite_jornada,
    validar_eventos_limite_jornada,
    validar_vacios_con_duracion_valida,
)


def _reutilizar_ids_conductor_exportacion(
    turnos_seleccionados: List[Dict[str, Any]],
    todos_eventos: List[Dict[str, Any]],
    config: Dict[str, Any],
) -> Tuple[List[Dict[str, Any]], Optional[Dict[str, int]]]:
    """
    Reutiliza IDs lógicos de conductor para exportación, manteniendo jornadas separadas.
    No altera turnos ni límites; solo renumera IDs en eventos para reducir cardinalidad.
    """
    f3_cfg = (config or {}).get("fase_3_union_conductores", {}) or {}
    n_logicos = len(turnos_seleccionados or [])
    if n_logicos <= 0:
        return todos_eventos, None

    descanso_min = int(f3_cfg.get("descanso_min_reutilizacion_min", 480) or 480)
    max_jornadas = int(f3_cfg.get("max_jornadas_por_conductor_exportacion", 2) or 2)

    linea_grupo: Dict[str, str] = {}
    for grupo, lineas in ((config or {}).get("grupos_lineas") or {}).items():
        for linea in (lineas or []):
            ls = str(linea).strip()
            if ls:
                linea_grupo[ls] = str(grupo).strip()
    grupos_por_conductor: Dict[int, frozenset] = {}
    for ev in (todos_eventos or []):
        if str(ev.get("evento", "")).strip().upper() != "COMERCIAL":
            continue
        cid = ev.get("conductor")
        try:
            cid_int = int(cid)
        except Exception:
            continue
        linea = str(ev.get("linea", "") or "").strip()
        grupo = linea_grupo.get(linea, "__SIN_GRUPO__")
        grupos_por_conductor.setdefault(cid_int, set()).add(grupo)
    grupos_por_conductor = {k: frozenset(v) for k, v in grupos_por_conductor.items()}

    def _to_min_local(v: Any) -> int:
        s = str(v or "").strip()
        if ":" in s:
            try:
                h, m = s.split(":")
                return int(h) * 60 + int(m)
            except Exception:
                return 0
        try:
            return int(float(s))
        except Exception:
            return 0

    intervalos_por_cid: Dict[int, Tuple[int, int]] = {}
    for ev in (todos_eventos or []):
        cid = ev.get("conductor")
        try:
            cid_int = int(cid)
        except Exception:
            continue
        ini_ev = _to_min_local(ev.get("inicio", 0))
        fin_ev = _to_min_local(ev.get("fin", 0))
        if fin_ev < ini_ev:
            fin_ev += 1440
        prev = intervalos_por_cid.get(cid_int)
        if prev is None:
            intervalos_por_cid[cid_int] = (ini_ev, fin_ev)
        else:
            intervalos_por_cid[cid_int] = (min(prev[0], ini_ev), max(prev[1], fin_ev))

    intervalos: List[Tuple[int, int, int]] = [
        (ini, fin, cid) for cid, (ini, fin) in intervalos_por_cid.items()
    ]
    if len(intervalos) < n_logicos:
        # Fallback seguro para IDs sin eventos (debería ser raro).
        for idx, t in enumerate(turnos_seleccionados, start=1):
            if idx in intervalos_por_cid:
                continue
            ini = int(t.get("inicio", 0) or 0)
            fin = int(t.get("fin", 0) or 0)
            if fin < ini:
                fin += 1440
            intervalos.append((ini, fin, idx))
    intervalos.sort(key=lambda x: (x[0], x[1], x[2]))

    # Cada bucket representa un conductor físico reutilizable (mismo grupo de líneas).
    buckets: List[Dict[str, Any]] = []
    mapping: Dict[int, int] = {}
    next_cid = 1
    reusos_hechos = 0

    for ini, fin, logical_id in intervalos:
        candidato_idx = -1
        mejor_fin = -1
        grupos_actual = grupos_por_conductor.get(logical_id, frozenset({"__SIN_GRUPO__"}))
        for i, b in enumerate(buckets):
            if int(b["jornadas"]) >= max_jornadas:
                continue
            if b.get("grupos") != grupos_actual:
                continue
            fin_prev = int(b["fin"])
            if fin_prev + descanso_min <= ini and fin_prev > mejor_fin:
                mejor_fin = fin_prev
                candidato_idx = i
        if candidato_idx >= 0:
            b = buckets[candidato_idx]
            b["fin"] = fin
            b["jornadas"] = int(b["jornadas"]) + 1
            cid = int(b["id"])
            reusos_hechos += 1
        else:
            cid = next_cid
            next_cid += 1
            buckets.append({"id": cid, "fin": fin, "jornadas": 1, "grupos": grupos_actual})
        mapping[logical_id] = cid

    eventos_remap: List[Dict[str, Any]] = []
    for ev in (todos_eventos or []):
        ev2 = dict(ev)
        cid_old = ev2.get("conductor")
        try:
            cid_int = int(cid_old)
        except Exception:
            eventos_remap.append(ev2)
            continue
        if cid_int in mapping:
            ev2["conductor"] = mapping[cid_int]
        eventos_remap.append(ev2)

    stats = {
        "logicos": n_logicos,
        "fisicos": len(set(mapping.values())),
        "reusos": reusos_hechos,
    }
    return eventos_remap, stats


def _auditar_excel_resultado(path_xlsx: str, config: Dict[str, Any]) -> None:
    """
    Auditoría final obligatoria sobre el Excel exportado.
    Reglas duras:
    - Sin solapes ni huecos por conductor (orden escrito en hoja).
    - Sin teletransportes (nodo destino previo == nodo origen siguiente, canónico).
    - Primer evento InS y último FnS por conductor.
    - InS/FnS siempre en depósito.
    - Sin exceso de jornada según límite global/por grupo.
    - Sin paradas consecutivas en EventosCompletos ni BusEventos.
    - Consistencia de inicio/fin de TurnosConductores vs EventosCompletos.
    """
    from collections import defaultdict, Counter
    from openpyxl import load_workbook

    def _to_min(v: Any) -> int:
        s = str(v or "").strip()
        if ":" in s:
            try:
                h, m = s.split(":")
                return int(h) * 60 + int(m)
            except Exception:
                return 0
        try:
            return int(float(s))
        except Exception:
            return 0

    def _dur(inicio: int, fin: int) -> int:
        d = int(fin) - int(inicio)
        if d < 0:
            d += 1440
        return d

    def _norm(s: Any) -> str:
        return str(s or "").strip().upper()

    def _canon_node(s: Any) -> str:
        return " ".join(_norm(s).replace("DEPOSITO", "").split())

    wb = load_workbook(path_xlsx, data_only=True)
    ws_ec = wb["EventosCompletos"]
    ws_tc = wb["TurnosConductores"]
    ws_be = wb["BusEventos"]

    linea_grupo: Dict[str, str] = {}
    for g, lineas in (config.get("grupos_lineas") or {}).items():
        for l in (lineas or []):
            linea_grupo[str(l).strip()] = str(g).strip()
    lim_global = int(config.get("limite_jornada", 600) or 600)
    lim_por = {str(k): int(v) for k, v in (config.get("limite_jornada_por_grupo_linea") or {}).items()}

    por_c: Dict[str, List[Tuple[Any, ...]]] = defaultdict(list)
    for r in ws_ec.iter_rows(min_row=2, values_only=True):
        c = str(r[2] or "").strip()
        if c:
            por_c[c].append(r)

    # Orden estable por conductor para evitar falsos positivos cuando el Excel
    # no viene estrictamente ordenado por tiempo/tipo.
    def _tipo_rank(tp: Any) -> int:
        t = _norm(tp)
        if t == "INS":
            return 0
        if t == "FNS":
            return 3
        if t == "PARADA":
            return 2
        return 1

    def _orden_evento_row(r: Tuple[Any, ...]) -> Tuple[int, int, int]:
        ini = _to_min(r[3])
        fin = _to_min(r[4])
        return (ini, fin, _tipo_rank(r[0]))

    solapes: List[Any] = []
    huecos: List[Any] = []
    tele: List[Any] = []
    first_last_bad: List[Any] = []
    ins_fns_bad: List[Any] = []
    jornada_bad: List[Any] = []

    for c, rows in por_c.items():
        rows = sorted(rows, key=_orden_evento_row)
        if not rows:
            continue
        if _norm(rows[0][0]) != "INS" or _norm(rows[-1][0]) != "FNS":
            first_last_bad.append((c, rows[0][0], rows[-1][0]))

        for rr in rows:
            tp = _norm(rr[0])
            if tp in ("INS", "FNS"):
                if "DEPOSITO" not in _norm(rr[6]) or "DEPOSITO" not in _norm(rr[7]):
                    ins_fns_bad.append((c, tp, rr[3], rr[4], rr[6], rr[7]))

        for i in range(len(rows) - 1):
            a, b = rows[i], rows[i + 1]
            tipo_a = _norm(a[0])
            tipo_b = _norm(b[0])
            fin_a = _to_min(a[4])
            ini_b = _to_min(b[3])
            if tipo_a == "FNS" and tipo_b == "INS":
                # Se permite hueco entre jornadas del mismo conductor físico.
                continue
            if ini_b < fin_a:
                solapes.append((c, a[0], b[0], a[4], b[3]))
            if ini_b > fin_a:
                huecos.append((c, a[0], b[0], a[4], b[3], ini_b - fin_a))
            if _canon_node(a[7]) != _canon_node(b[6]):
                tele.append((c, a[0], b[0], a[7], b[6], a[4], b[3]))

        # Validar límite por jornada (cada bloque InS -> FnS), no sobre el span total.
        ini_j = None
        lineas_j: List[str] = []
        for r in rows:
            tp = _norm(r[0])
            if tp == "INS":
                ini_j = _to_min(r[3])
                lineas_j = []
                continue
            if tp == "COMERCIAL":
                linea_r = str(r[9] or "").strip()
                if linea_r:
                    lineas_j.append(linea_r)
                continue
            if tp == "FNS" and ini_j is not None:
                fin_j = _to_min(r[4])
                jornada = _dur(ini_j, fin_j)
                grp = ""
                if lineas_j:
                    grp = linea_grupo.get(Counter(lineas_j).most_common(1)[0][0], "")
                lim = lim_por.get(grp, lim_global)
                if jornada > lim:
                    jornada_bad.append((c, grp, lim, jornada, jornada - lim))
                ini_j = None
                lineas_j = []

    ins_mismatch: List[Any] = []
    fns_mismatch: List[Any] = []
    tc = {
        str(r[0]).strip(): r
        for r in ws_tc.iter_rows(min_row=2, values_only=True)
        if str(r[0] or "").strip()
    }
    for c, rows in por_c.items():
        rows = sorted(rows, key=_orden_evento_row)
        if c not in tc:
            continue
        ins = [_to_min(r[3]) for r in rows if _norm(r[0]) == "INS"]
        fns = [_to_min(r[4]) for r in rows if _norm(r[0]) == "FNS"]
        if ins and _to_min(tc[c][3]) != min(ins):
            ins_mismatch.append((c, _to_min(tc[c][3]), min(ins)))
        if fns and _to_min(tc[c][4]) != max(fns):
            fns_mismatch.append((c, _to_min(tc[c][4]), max(fns)))

    paradas_ec: List[Any] = []
    parada_post_ins: List[Any] = []
    parada_pre_fns: List[Any] = []
    for c, rows in por_c.items():
        rows = sorted(rows, key=_orden_evento_row)
        for i in range(len(rows) - 1):
            if _norm(rows[i][0]) == "PARADA" and _norm(rows[i + 1][0]) == "PARADA":
                paradas_ec.append((c, rows[i][3], rows[i][4], rows[i + 1][3], rows[i + 1][4]))
            if _norm(rows[i][0]) == "INS" and _norm(rows[i + 1][0]) == "PARADA":
                parada_post_ins.append((c, rows[i][4], rows[i + 1][3], rows[i + 1][4]))
            if _norm(rows[i][0]) == "PARADA" and _norm(rows[i + 1][0]) == "FNS":
                parada_pre_fns.append((c, rows[i][3], rows[i][4], rows[i + 1][3]))

    por_b: Dict[str, List[Tuple[Any, ...]]] = defaultdict(list)
    for r in ws_be.iter_rows(min_row=2, values_only=True):
        b = str(r[8] or "").strip()
        if b:
            por_b[b].append(r)
    paradas_be: List[Any] = []
    buses_inicio_fin_no_deposito: List[Any] = []
    dep_canon = _canon_node(config.get("deposito", ""))
    for b, rows in por_b.items():
        rows_ord = sorted(rows, key=lambda rr: (_to_min(rr[2]), _to_min(rr[4])))
        if rows_ord:
            first = rows_ord[0]
            last = rows_ord[-1]
            # REGLA DURA: cada bus debe iniciar en depósito y terminar en depósito.
            if _canon_node(first[3]) != dep_canon or _canon_node(last[5]) != dep_canon:
                buses_inicio_fin_no_deposito.append((b, first[3], first[2], last[5], last[4]))
        for i in range(len(rows_ord) - 1):
            if _norm(rows_ord[i][1]) == "PARADA" and _norm(rows_ord[i + 1][1]) == "PARADA":
                paradas_be.append((b, rows_ord[i][2], rows_ord[i][4], rows_ord[i + 1][2], rows_ord[i + 1][4]))

    violaciones = [
        ("solapes", solapes),
        ("huecos", huecos),
        ("teletransportes", tele),
        ("primer_ultimo_evento", first_last_bad),
        ("ins_fns_fuera_deposito", ins_fns_bad),
        ("jornada_exceso", jornada_bad),
        ("ins_mismatch_tc_ec", ins_mismatch),
        ("fns_mismatch_tc_ec", fns_mismatch),
        ("paradas_consecutivas_ec", paradas_ec),
        ("parada_despues_ins", parada_post_ins),
        ("parada_antes_fns", parada_pre_fns),
        ("bus_inicio_fin_no_deposito", buses_inicio_fin_no_deposito),
        ("paradas_consecutivas_be", paradas_be),
    ]
    fallas = [(k, v) for k, v in violaciones if v]
    if fallas:
        lineas = ["[AUDITORIA FINAL - REGLA DURA] Se detectaron inconsistencias en el Excel exportado:"]
        for k, v in fallas:
            lineas.append(f"  - {k}: {len(v)} (ej: {v[:3]})")
        raise ValueError("\n".join(lineas))

    print("  [OK] Auditoría final integral: sin solapes, sin huecos, sin teletransportes, "
          "InS/FnS consistentes, jornadas OK y sin paradas consecutivas.")


def main(
    archivo_excel: str = "datos_salidas.xlsx",
    archivo_config: str = "configuracion.json",
    archivo_salida: str = "resultado_diagramacion.xlsx",
    random_seed: Optional[int] = 42,
) -> None:
    """
    Punto de entrada principal. Orquesta la carga de datos, ambas fases de optimización
    y la exportación final reutilizando un único GestorDeLogistica.
    """
    # Detectar si estamos ejecutando desde un ejecutable compilado
    es_ejecutable = getattr(sys, 'frozen', False)
    
    if es_ejecutable:
        # Si estamos en un ejecutable, usar el directorio de trabajo actual
        # (donde está el .exe) en lugar del directorio del paquete
        raiz_proyecto = os.getcwd()
    else:
        # Si estamos ejecutando desde Python normal, usar la lógica original
        paquete_dir = os.path.abspath(os.path.dirname(__file__))
        raiz_proyecto = os.path.abspath(os.path.join(paquete_dir, os.pardir))
    
    # Si las rutas ya son absolutas, usarlas directamente; si no, unirlas con raiz_proyecto
    if os.path.isabs(archivo_config):
        ruta_config = archivo_config
    else:
        ruta_config = os.path.join(raiz_proyecto, archivo_config)
    
    if os.path.isabs(archivo_excel):
        ruta_excel = archivo_excel
    else:
        ruta_excel = os.path.join(raiz_proyecto, archivo_excel)
    
    if os.path.isabs(archivo_salida):
        ruta_salida = archivo_salida
    else:
        ruta_salida = os.path.join(raiz_proyecto, archivo_salida)
    
    # Asegurar que las rutas sean absolutas
    ruta_config = os.path.abspath(ruta_config)
    ruta_excel = os.path.abspath(ruta_excel)
    ruta_salida = os.path.abspath(ruta_salida)

    print("\n" + "=" * 70)
    print("ARCHIVOS DE ESTA EJECUCIÓN")
    print("=" * 70)
    print(f"  Entrada:  {ruta_excel}")
    print(f"  Salida:   {ruta_salida}")
    print("  Cierra Excel si tienes el archivo de salida abierto para que se pueda guardar.")
    print("=" * 70 + "\n")

    if not os.path.exists(ruta_config):
        print(f"ERROR: No se encontró el archivo de configuración en {ruta_config}")
        return
    if not os.path.exists(ruta_excel):
        print(f"ERROR: No se encontró el archivo Excel en {ruta_excel}")
        return

    try:
        config = cargar_config(ruta_config)
        config = autocompletar_configuracion(config)
    except Exception as e:
        print(f"ERROR: No se pudo cargar configuración de forma estricta. {e}")
        return
    try:
        validar_configuracion(config)
    except ConfigValidationError as e:
        print(f"ERROR: Configuración inválida. {e}")
        return
    try:
        viajes = cargar_salidas_desde_excel(ruta_excel)
    except Exception as e:
        print(f"ERROR: Excel de viajes inválido en modo estricto. {e}")
        return
    if not viajes:
        print("ERROR: No se encontraron viajes comerciales en el Excel.")
        return

    gestor = GestorDeLogistica(config)
    modo_verbose = bool(config.get("modo_verbose", False))
    opt_iter = config.get("optimizacion_iterativa", {}) or {}
    max_iter = int(opt_iter.get("max_iteraciones", 1))
    # Por defecto NO repetir búsquedas: cortar en la primera solución válida.
    buscar_mejor_solucion = bool(opt_iter.get("buscar_mejor_solucion", False))
    max_reintentos_jornada = int(opt_iter.get("max_reintentos_jornada", 3))
    if not buscar_mejor_solucion:
        max_iter = 1
        max_reintentos_jornada = 0
    total_intentos = max_iter + max(0, max_reintentos_jornada)
    ultima_firma_error = ""
    repeticiones_mismo_error = 0

    mejor_turnos = None
    mejor_bloques = None
    mejor_eventos_bus = None
    mejor_metadata = None
    mejor_status_f1 = mejor_status_f2 = mejor_status_f3 = ""
    mejor_conteo = 999999

    for iteracion in range(total_intentos):
        seed_actual = (random_seed or 42) + iteracion * 1000
        if iteracion < max_iter and max_iter > 1:
            print(f"\n--- Iteración {iteracion + 1}/{max_iter} (seed={seed_actual}) ---")
        elif iteracion >= max_iter:
            extra_idx = iteracion - max_iter + 1
            print(f"\n--- Reintento automático por validez de jornada {extra_idx}/{max_reintentos_jornada} (seed={seed_actual}) ---")

        # Flujo: Fase 1 -> Fase 2 -> Fase 3.
        try:
            print("Iniciando Fase 1 (Buses) ...")
            bloques_bus, eventos_bus, status_f1 = resolver_diagramacion_buses(
                config,
                viajes,
                gestor,
                random_seed=seed_actual,
                verbose=modo_verbose,
            )
            if not bloques_bus:
                print("No se generaron bloques de buses.")
                continue
            ids_fase1 = set()
            for bloque in bloques_bus:
                for ev in bloque:
                    if isinstance(ev, dict) and "id" in ev:
                        ids_fase1.add(ev["id"])
            ids_viajes = {v["id"] for v in viajes}
            if ids_fase1 != ids_viajes and iteracion == 0:
                faltan = ids_viajes - ids_fase1
                if faltan:
                    print(f"  [AVISO] Viajes sin asignar en Fase 1: {len(faltan)}")
            elif ids_fase1 == ids_viajes and iteracion == 0:
                print(f"  [OK] Mismos viajes: {len(ids_viajes)} en Fase 1.")
        except Exception as e:
            print(f"[ERROR] Error en Fase 1: {e}")
            if iteracion == 0:
                import traceback
                traceback.print_exc()
            continue

        try:
            print("Iniciando Fase 2 (Conductores) ...")
            turnos_seleccionados, metadata_tareas, status_f2 = resolver_diagramacion_conductores(
                config,
                viajes,
                bloques_bus,
                gestor,
                verbose=modo_verbose,
            )
        except Exception as e:
            print(f"[ERROR] Error en Fase 2: {e}")
            if iteracion == 0:
                import traceback
                traceback.print_exc()
            continue

        try:
            print("Iniciando Fase 3 (Unión de Conductores) ...")
            n_turnos_fase2 = len(turnos_seleccionados or [])
            turnos_seleccionados, status_f3 = resolver_union_conductores(
                config,
                turnos_seleccionados,
                metadata_tareas,
                viajes,
                gestor,
                verbose=modo_verbose,
                seed_externo=seed_actual,
            )
            n_turnos_fase3 = len(turnos_seleccionados or [])
            if n_turnos_fase3 > n_turnos_fase2:
                raise ValueError(
                    f"[FASE 3 - REGLA DURA] Fase 3 aumentó turnos: {n_turnos_fase2} -> {n_turnos_fase3}. "
                    "Fase 3 solo puede mantener o reducir."
                )
            # REGLA DURA: no aceptar combinaciones que superen límite jornada.
            validar_turnos_limite_jornada(turnos_seleccionados, gestor.limite_jornada)
        except Exception as e:
            print(f"[ERROR] Error en Fase 3: {e}")
            firma = str(e).strip()
            if firma and firma == ultima_firma_error:
                repeticiones_mismo_error += 1
            else:
                ultima_firma_error = firma
                repeticiones_mismo_error = 1
            if repeticiones_mismo_error >= 3:
                print("[ERROR] Mismo error de validez repetido 3 veces. Se detiene para evitar bucle.")
                break
            if iteracion == 0:
                import traceback
                traceback.print_exc()
            continue

        conteo = len(turnos_seleccionados) if turnos_seleccionados else 999999
        if conteo < mejor_conteo:
            mejor_conteo = conteo
            mejor_turnos = turnos_seleccionados
            mejor_bloques = bloques_bus
            mejor_eventos_bus = eventos_bus
            mejor_metadata = metadata_tareas
            mejor_status_f1, mejor_status_f2, mejor_status_f3 = status_f1, status_f2, status_f3
            if max_iter > 1:
                print(f"  [OK] Nueva mejor solución: {conteo} conductores")
        # Modo normal: cortar en la primera solución válida para evitar ciclos.
        # Modo búsqueda de mejor solución: correr hasta max_iter y luego cortar.
        if mejor_turnos is not None and (not buscar_mejor_solucion or iteracion >= (max_iter - 1)):
            break

    if mejor_turnos is None:
        print(
            "[ERROR] No se pudo generar ninguna solución válida sin violar límites de jornada "
            f"tras {total_intentos} intentos."
        )
        return

    turnos_seleccionados = mejor_turnos
    n_turnos_logicos_fase3 = len(turnos_seleccionados or [])
    bloques_bus = mejor_bloques
    eventos_bus = mejor_eventos_bus
    metadata_tareas = mejor_metadata
    status_f1, status_f2, status_f3 = mejor_status_f1, mejor_status_f2, mejor_status_f3
    # La factibilización operativa debe venir resuelta desde Fase 2/Fase 3.
    # Main no debe capar ni modificar turnos para exportación.

    ruta_salida_abs = os.path.abspath(ruta_salida)
    directorio_salida = os.path.dirname(ruta_salida_abs)
    
    # Asegurar que el directorio de salida existe
    if directorio_salida and not os.path.exists(directorio_salida):
        try:
            os.makedirs(directorio_salida, exist_ok=True)
            print(f"[OK] Directorio de salida creado: {directorio_salida}")
        except Exception as e_dir:
            print(f"[ERROR] No se pudo crear el directorio de salida: {directorio_salida}")
            print(f"  Error: {e_dir}")
            return
    
    print(f"\n{'=' * 80}")
    print(f"EXPORTANDO RESULTADOS")
    print(f"{'=' * 80}")
    print(f"  Ruta de salida: {ruta_salida_abs}")
    print(f"  Directorio: {directorio_salida}")
    print(f"  Directorio existe: {os.path.exists(directorio_salida) if directorio_salida else 'N/A'}")
    print(f"{'=' * 80}\n")
    
    # Validar datos antes de exportar
    print(f"\n{'=' * 80}")
    print(f"VALIDACIÓN ANTES DE EXPORTAR")
    print(f"{'=' * 80}")
    print(f"  Bloques de buses: {len(bloques_bus) if bloques_bus else 0}")
    print(f"  Turnos recibidos (Fase 3): {len(turnos_seleccionados) if turnos_seleccionados else 0}")
    print(f"  Viajes comerciales: {len(viajes) if viajes else 0}")
    print(f"  Ruta de salida: {ruta_salida_abs}")
    print(f"  Directorio existe: {os.path.exists(directorio_salida) if directorio_salida else False}")
    print(f"{'=' * 80}\n")
    
    if not bloques_bus or len(bloques_bus) == 0:
        print("  [ERROR] No hay bloques de buses para exportar!")
        return
    
    if not turnos_seleccionados or len(turnos_seleccionados) == 0:
        print("  [ERROR] No hay turnos seleccionados para exportar!")
        return

    # Filtro final: no exportar conductores sin tareas (InS/FnS solos = innecesarios).
    # Considerar comerciales por id, _tmp_id y por tareas en metadata (ej. _ev_bus_idx) para 100% cobertura.
    ids_comerciales = set()
    for v in (viajes or []):
        for key in (v.get("id"), v.get("_tmp_id")):
            if key is not None:
                ids_comerciales.add(key)
                ids_comerciales.add(str(key))
    # Incluir tids que aparecen en metadata_tareas (viajes de bloques con id sintético _ev_*)
    for tid in (metadata_tareas or {}):
        if tid is not None:
            ids_comerciales.add(tid)
            ids_comerciales.add(str(tid))
    turnos_con_comerciales = [
        t for t in turnos_seleccionados
        if any(
            tid in ids_comerciales or str(tid) in ids_comerciales
            for tid, _ in t.get("tareas_con_bus", [])
        )
    ]
    eliminados = len(turnos_seleccionados) - len(turnos_con_comerciales)
    if eliminados > 0:
        print(f"  [FILTRO] Excluidos {eliminados} conductores sin eventos comerciales (no exportados)")
        turnos_seleccionados = turnos_con_comerciales

    limites_por_conductor = {
        idx: int(t.get("limite_jornada_aplicable", gestor.limite_jornada) or gestor.limite_jornada)
        for idx, t in enumerate(turnos_seleccionados, start=1)
    }

    # REGLA DURA: Ningún conductor puede superar el límite máximo de jornada.
    validar_turnos_limite_jornada(turnos_seleccionados, gestor.limite_jornada)
    print("  [OK] Validación límite jornada: ningún conductor supera el máximo.")
    eventos_bus_planos = [ev for bloque in (eventos_bus or []) for ev in (bloque or [])]
    validar_vacios_con_duracion_valida(eventos_bus_planos)
    print("  [OK] Validación vacíos (BusEventos): duración válida entre nodos.")

    # Construir eventos completos como parte del cálculo (Fase 2/3); el exportador solo escribe.
    todos_eventos = construir_eventos_completos(
        gestor,
        bloques_bus,
        turnos_seleccionados,
        viajes,
        metadata_tareas,
        eventos_bus=eventos_bus,
        limites_por_conductor=limites_por_conductor,
    )
    # REGLA EN EXTREMO DURA: validar que ningún conductor supere el límite en eventos.
    validar_eventos_limite_jornada(todos_eventos, gestor.limite_jornada, limites_por_conductor=limites_por_conductor)
    validar_vacios_con_duracion_valida(todos_eventos)
    print("  [OK] Validación vacíos (EventosCompletos): duración válida entre nodos.")

    conductores_exportados = None
    ruta_export_real = ruta_salida_abs
    try:
        print("Iniciando exportación de resultados...")
        resultado_export = exportar_resultado_excel(
            config,
            bloques_bus,
            turnos_seleccionados,
            viajes,
            metadata_tareas,
            status_f1,
            status_f2,
            ruta_salida_abs,
            gestor=gestor,
            verbose=modo_verbose,
            status_f3=status_f3,
            eventos_bus=eventos_bus,
            eventos_completos=todos_eventos,
        )
        print(f"[OK] Exportacion completada: {ruta_salida_abs}")
        conductores_exportados = (
            resultado_export.get("conductores_exportados")
            if isinstance(resultado_export, dict) else None
        )
        if isinstance(resultado_export, dict):
            ruta_export_real = resultado_export.get("path_real_guardado", ruta_salida_abs) or ruta_salida_abs
        if ruta_export_real != ruta_salida_abs:
            print(f"[INFO] Archivo exportado real: {ruta_export_real}")
    except Exception as e:
        print(f"  [ERROR] Error en exportacion: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    # Verificar que el archivo se generó correctamente
    if os.path.exists(ruta_export_real):
        tamaño = os.path.getsize(ruta_export_real)
        print(f"\n{'=' * 80}")
        print(f"VERIFICACIÓN FINAL DEL ARCHIVO GENERADO")
        print(f"{'=' * 80}")
        print(f"  Archivo: {ruta_export_real}")
        print(f"  Tamaño: {tamaño:,} bytes ({tamaño / 1024:.2f} KB)")
        print(f"  Existe: SÍ")
        print(f"{'=' * 80}\n")
        _auditar_excel_resultado(ruta_export_real, config)
    else:
        print(f"\n{'=' * 80}")
        print(f"[ADVERTENCIA] EL ARCHIVO NO SE GENERO")
        print(f"{'=' * 80}")
        print(f"  Ruta esperada: {ruta_salida_abs}")
        print(f"  Directorio actual: {os.getcwd()}")
        print(f"{'=' * 80}\n")

    print("\n" + "=" * 80)
    print("--- RESUMEN FINAL DEL DIAGRAMADOR ---")
    print("=" * 80)
    print(f"Total de viajes comerciales procesados: {len(viajes)}")
    print(f"Total de buses utilizados (bloques): {len(bloques_bus)}")
    n_fase3 = n_turnos_logicos_fase3
    n_export = conductores_exportados if conductores_exportados is not None else n_fase3
    print(f"Total de conductores lógicos (salida Fase 3): {n_fase3}")
    print(f"Total de segmentos operativos exportados (TurnosConductores/EventosCompletos): {n_export}")
    print(f"Estado Optimización Fase 1 (Buses): {status_f1}")
    print(f"Estado Optimización Fase 2 (Conductores): {status_f2}")
    print(f"Estado Optimización Fase 3 (unión pura): {status_f3}")
    if n_fase3 != 0:
        print(f"Estado Fase 3 oficial (factible operativo, usado en exportación): {n_fase3} turnos")
    print("=" * 80)


if __name__ == "__main__":
    main()

