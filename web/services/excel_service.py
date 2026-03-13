"""
Servicios para manejo de archivos Excel.
Incluye procesamiento de viajes comerciales y configuración.
"""

import io
import os
import copy
import tempfile
import shutil
import time
import pandas as pd
import traceback
from typing import Any, Dict, List, Optional, Tuple
from werkzeug.datastructures import FileStorage
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .config_service import ConfigService, TIPOS_BUS_DISPONIBLES
from ..utils.logging_utils import logger
from ..utils.validators import validate_excel_file


def _convertir_habilitado_a_booleano(valor: Any, valor_por_defecto: bool = True) -> bool:
    """
    Convierte un valor a booleano para campos 'habilitado'.
    
    Acepta múltiples formatos:
    - Texto: "SI", "SÍ", "YES", "TRUE", "VERDADERO" → True
    - Texto: "NO", "FALSE", "FALSO" → False
    - Booleanos: True/False
    - Números: 1, "1" → True; 0, "0" → False
    - None o vacío: usa valor_por_defecto
    
    REGLA IMPORTANTE: Si el valor es "NO", siempre retorna False, incluso si está vacío.
    
    Args:
        valor: Valor a convertir (puede ser texto, booleano, número, None)
        valor_por_defecto: Valor a usar si el valor está vacío o no es reconocible
        
    Returns:
        Valor booleano
    """
    # Manejar None primero
    if valor is None:
        return valor_por_defecto
    
    # Si es booleano, retornarlo directamente
    if isinstance(valor, bool):
        return valor
    
    # Si es número, convertir a booleano (0 = False, cualquier otro = True)
    if isinstance(valor, (int, float)):
        return bool(valor)
    
    # Convertir a string y limpiar
    # IMPORTANTE: No usar strip() todavía, primero verificar si es string vacío
    valor_str_original = str(valor)
    valor_str = valor_str_original.strip().upper()
    
    # Si después de limpiar está vacío, usar valor por defecto
    if not valor_str or valor_str in ("", "NONE", "NULL", "NAN"):
        return valor_por_defecto
    
    # IMPORTANTE: Verificar valores FALSE primero para tener prioridad
    # Valores que indican NO habilitado (False) - PRIORIDAD ALTA
    valores_false = {"NO", "FALSE", "FALSO", "0", "F", "N", "DESHABILITADO", "DISABLED"}
    if valor_str in valores_false:
        logger.debug(f"Valor '{valor_str_original}' interpretado como NO habilitado (False)")
        return False
    
    # Valores que indican habilitado (True)
    valores_true = {"SI", "SÍ", "YES", "TRUE", "VERDADERO", "1", "V", "S", "HABILITADO", "ENABLED"}
    if valor_str in valores_true:
        logger.debug(f"Valor '{valor_str_original}' interpretado como habilitado (True)")
        return True
    
    # Si no es reconocible, usar valor por defecto y loggear advertencia
    logger.warning(
        f"Valor no reconocido para 'habilitado': '{valor_str_original}' (tipo: {type(valor).__name__}). "
        f"Usando valor por defecto: {valor_por_defecto}"
    )
    return valor_por_defecto


class ExcelTripsService:
    """
    Servicio para procesar archivos Excel de viajes comerciales.
    """
    
    STANDARD_FILE_NAME = "datos_salidas.xlsx"

    def __init__(
        self,
        config_service: ConfigService,
        trips_file_path: Optional[str] = None,
    ):
        """
        Inicializa el servicio de Excel de viajes.
        
        Args:
            config_service: Instancia de ConfigService para actualizar configuración
            trips_file_path: Ruta del archivo Excel de viajes (ej. diagramador_optimizado/datos_salidas.xlsx).
                Si no se indica, se usa el mismo directorio que configuracion.json con STANDARD_FILE_NAME.
        """
        self.config_service = config_service
        if trips_file_path is not None:
            self._trips_file_path = os.path.abspath(trips_file_path)
        else:
            config_dir = os.path.dirname(os.path.abspath(config_service.config_path))
            self._trips_file_path = os.path.abspath(os.path.join(config_dir, self.STANDARD_FILE_NAME))
    
    def process_upload(self, archivo: Optional[FileStorage]) -> Dict[str, Any]:
        """
        Procesa un archivo Excel subido y actualiza la configuración.
        
        Args:
            archivo: Archivo subido por el usuario
            
        Returns:
            Diccionario con resultado de la operación
        """
        # Validar archivo
        es_valido, mensaje_error = validate_excel_file(archivo)
        if not es_valido:
            return {"success": False, "message": mensaje_error}

        logger.info("process_upload: guardando SIEMPRE en datos_salidas.xlsx (sobrescritura, sin copias con fecha).")
        
        try:
            # Directorio de destino (ej. diagramador_optimizado/)
            dir_path = os.path.dirname(self._trips_file_path)
            if dir_path and not os.path.isdir(dir_path):
                os.makedirs(dir_path, exist_ok=True)

            # SIEMPRE un solo archivo: datos_salidas.xlsx. Sobrescribir; NUNCA crear copias con fecha.
            archivo_path = os.path.abspath(self._trips_file_path)

            # Eliminar archivo puntero de la lógica antigua (datos_salidas_actual o .txt) si existe
            for nombre_puntero in ("datos_salidas_actual.txt", "datos_salidas_actual"):
                ruta_puntero = os.path.join(dir_path or ".", nombre_puntero)
                try:
                    if os.path.isfile(ruta_puntero):
                        os.remove(ruta_puntero)
                        logger.info(f"Eliminado archivo puntero antiguo: {nombre_puntero}")
                except Exception:
                    pass

            # Eliminar TODAS las copias con timestamp (datos_salidas_YYYYMMDD_HHMMSS.xlsx)
            try:
                for nombre in os.listdir(dir_path or "."):
                    if nombre.startswith("datos_salidas_") and nombre.endswith(".xlsx"):
                        ruta_vieja = os.path.join(dir_path, nombre)
                        if os.path.isfile(ruta_vieja):
                            os.remove(ruta_vieja)
                            logger.info(f"Eliminada copia con fecha: {nombre}")
            except Exception as e_clean:
                logger.warning(f"No se pudieron limpiar copias antiguas: {e_clean}")

            fd_temp, temp_path = tempfile.mkstemp(suffix=".xlsx", dir=dir_path, prefix="upload_")
            try:
                os.close(fd_temp)
                archivo.save(temp_path)
                if hasattr(os, "sync"):
                    os.sync()
                # Sobrescribir siempre datos_salidas.xlsx (eliminar destino si existe para forzar reemplazo)
                if os.path.isfile(archivo_path):
                    try:
                        os.remove(archivo_path)
                    except PermissionError:
                        raise PermissionError(
                            "No se pudo guardar: cierra 'datos_salidas.xlsx' si lo tienes abierto en Excel y vuelve a subir."
                        )
                shutil.move(temp_path, archivo_path)
                logger.info(f"Archivo guardado correctamente en datos_salidas.xlsx (sobrescrito): {archivo_path}")
            except PermissionError:
                if os.path.exists(temp_path):
                    try:
                        os.remove(temp_path)
                    except Exception:
                        pass
                raise PermissionError(
                    "No se pudo guardar: cierra 'datos_salidas.xlsx' si lo tienes abierto en Excel y vuelve a subir."
                )
            except Exception as e_save:
                if os.path.exists(temp_path):
                    try:
                        os.remove(temp_path)
                    except Exception:
                        pass
                raise e_save
            finally:
                if os.path.exists(temp_path):
                    try:
                        os.remove(temp_path)
                    except Exception:
                        pass
            
            # Cargar configuración actual
            config = self.config_service.get_config()
            nodos_existentes_antes = set(config.get("nodos", []))
            
            # Procesar Excel y detectar nodos (lectura fresca desde disco)
            nodos_detectados = self._detect_nodes_from_excel(archivo_path)
            n_viajes = 0
            try:
                df_viajes = pd.read_excel(archivo_path)
                n_viajes = len(df_viajes) if not df_viajes.empty else 0
                logger.info(f"Archivo cargado: {n_viajes} filas de viajes en {archivo_path}")
            except Exception:
                pass

            if not nodos_detectados:
                return {
                    "success": False,
                    "message": "No se pudieron detectar nodos en el archivo",
                }
            
            # Actualizar configuración con nodos detectados
            self.config_service.update_nodes(nodos_detectados)
            self.config_service.save_config()
            
            # Calcular nodos nuevos
            nodos_nuevos = set(nodos_detectados) - nodos_existentes_antes
            depositos_activos = self.config_service.get_active_deposits()
            config_actualizada = self.config_service.get_config()
            
            # Construir mensaje
            if nodos_nuevos:
                mensaje = (
                    f"Archivo procesado correctamente. Se detectaron {len(nodos_detectados)} nodos "
                    f"({len(nodos_nuevos)} nuevos: {', '.join(sorted(nodos_nuevos))}). "
                    f"Todas las conexiones se regeneraron dinámicamente."
                )
            else:
                mensaje = (
                    f"Archivo procesado correctamente. Se detectaron {len(nodos_detectados)} nodos. "
                    f"Todas las conexiones se regeneraron dinámicamente."
                )
            if n_viajes > 0:
                mensaje += f" Se cargaron {n_viajes} viajes. "

            depositos_texto = ", ".join(depositos_activos)
            mensaje += (
                f" Depósitos activos: {depositos_texto}. "
                f"Conexiones: {len(config_actualizada.get('vacios', {}))} vacíos, "
                f"{len(config_actualizada.get('desplazamientos', {}))} desplazamientos. "
                f"El archivo se guardó en datos_salidas.xlsx y será usado para las optimizaciones."
            )
            
            return {
                "success": True,
                "message": mensaje,
                "nodos": nodos_detectados,
                "archivo_cargado": True,
                "nombre_archivo": self.STANDARD_FILE_NAME,
                "n_viajes": n_viajes,
            }
            
        except Exception as e:
            logger.error(f"Error procesando Excel: {e}")
            return {"success": False, "message": f"Error procesando archivo: {str(e)}"}
    
    def _detect_nodes_from_excel(self, archivo_path: str) -> List[str]:
        """
        Detecta nodos únicos desde un archivo Excel de viajes y, de forma
        adicional, actualiza la configuración con las líneas detectadas.
        
        Args:
            archivo_path: Ruta al archivo Excel
            
        Returns:
            Lista de nodos únicos detectados
        """
        try:
            # Verificar que el archivo existe
            if not os.path.exists(archivo_path):
                logger.error(f"El archivo no existe: {archivo_path}")
                return []
            
            # Leer el archivo Excel
            df = pd.read_excel(archivo_path)
            
            # Verificar que el DataFrame no esté vacío
            if df.empty:
                logger.error("El archivo Excel está vacío")
                return []
            
            # Verificar que tenga columnas
            if df.columns.empty:
                logger.error("El archivo Excel no tiene columnas")
                return []
            
            logger.info(f"Columnas encontradas en el archivo: {list(df.columns)}")
            
            # Buscar columnas de origen y destino de forma flexible (case-insensitive)
            # Primero intentar coincidencias exactas, luego parciales
            col_origen = None
            col_destino = None
            
            # Lista de posibles nombres para origen y destino (más completa)
            posibles_origen = ["origen", "origen_", "origen ", "origen_nodo", "nodo_origen", 
                             "desde", "from", "salida", "partida", "inicio"]
            posibles_destino = ["destino", "destino_", "destino ", "destino_nodo", "nodo_destino",
                              "hasta", "to", "llegada", "arribo", "fin"]
            
            # Primera pasada: búsqueda exacta
            for col in df.columns:
                col_lower = str(col).strip().lower()
                # Limpiar espacios y caracteres especiales
                col_clean = col_lower.replace("_", "").replace("-", "").replace(" ", "")
                
                if col_clean in [p.replace("_", "").replace("-", "").replace(" ", "") for p in posibles_origen]:
                    col_origen = col
                    logger.info(f"Columna de origen encontrada (exacta): '{col}'")
                    break
                elif col_clean in [p.replace("_", "").replace("-", "").replace(" ", "") for p in posibles_destino]:
                    col_destino = col
                    logger.info(f"Columna de destino encontrada (exacta): '{col}'")
                    break
            
            # Segunda pasada: búsqueda parcial (contains)
            if not col_origen or not col_destino:
                for col in df.columns:
                    col_lower = str(col).strip().lower()
                    if not col_origen and any(p in col_lower for p in posibles_origen):
                        col_origen = col
                        logger.info(f"Columna de origen encontrada (parcial): '{col}'")
                    if not col_destino and any(p in col_lower for p in posibles_destino):
                        col_destino = col
                        logger.info(f"Columna de destino encontrada (parcial): '{col}'")
            
            # Si no se encontraron, intentar con las primeras dos columnas si hay al menos 2
            if not col_origen:
                if len(df.columns) >= 1:
                    # Intentar inferir: primera columna puede ser origen
                    col_origen = df.columns[0]
                    logger.warning(f"No se encontró columna 'Origen', usando '{col_origen}' como origen")
                else:
                    logger.error("No se encontró columna de origen y el archivo no tiene suficientes columnas")
            
            if not col_destino:
                if len(df.columns) >= 2:
                    # Intentar inferir: segunda columna puede ser destino
                    col_destino = df.columns[1]
                    logger.warning(f"No se encontró columna 'Destino', usando '{col_destino}' como destino")
                elif len(df.columns) >= 1 and col_origen != df.columns[0]:
                    # Si solo hay una columna y no es la de origen, usarla como destino
                    col_destino = df.columns[0]
                    logger.warning(f"No se encontró columna 'Destino', usando '{col_destino}' como destino")
                else:
                    logger.error("No se encontró columna de destino y el archivo no tiene suficientes columnas")
            
            # Validar que al menos una columna fue encontrada
            if not col_origen and not col_destino:
                logger.error("No se pudieron identificar columnas de origen ni destino en el archivo")
                logger.info(f"Columnas disponibles: {list(df.columns)}")
                logger.info(f"Primeras 5 filas del DataFrame:\n{df.head()}")
                return []
            
            # Detectar nodos únicos
            nodos_detectados = []
            
            if col_origen:
                try:
                    nodos_origen = df[col_origen].dropna().unique()
                    nodos_detectados.extend([str(n).strip() for n in nodos_origen if str(n).strip()])
                    logger.info(f"Detectados {len(nodos_origen)} nodos desde columna '{col_origen}'")
                except Exception as e:
                    logger.error(f"Error leyendo columna de origen '{col_origen}': {e}")
            
            if col_destino:
                try:
                    nodos_destino = df[col_destino].dropna().unique()
                    nodos_detectados.extend([str(n).strip() for n in nodos_destino if str(n).strip()])
                    logger.info(f"Detectados {len(nodos_destino)} nodos desde columna '{col_destino}'")
                except Exception as e:
                    logger.error(f"Error leyendo columna de destino '{col_destino}': {e}")
            
            # Eliminar duplicados y valores vacíos
            nodos_detectados = sorted(list(set([n for n in nodos_detectados if n])))
            
            if not nodos_detectados:
                logger.error("No se detectaron nodos válidos en el archivo")
                logger.info(f"Columnas disponibles: {list(df.columns)}")
                logger.info(f"Primeras filas del archivo:\n{df.head()}")
            
            # Detectar líneas únicas desde la columna "Linea" (si existe)
            lineas_detectadas: List[str] = []
            col_linea = None
            
            # Buscar columna de línea de forma flexible
            for col in df.columns:
                col_lower = str(col).strip().lower()
                if col_lower in ["linea", "línea", "linea_", "línea_", "linea_nombre", "codigo_linea"]:
                    col_linea = col
                    break
            
            if col_linea:
                try:
                    lineas_raw = df[col_linea].dropna().unique()
                    lineas_detectadas = [
                        str(linea).strip()
                        for linea in lineas_raw
                        if str(linea).strip() != ""
                    ]
                    logger.info(f"Detectadas {len(lineas_detectadas)} líneas desde columna '{col_linea}'")
                except Exception as e:
                    logger.warning(f"Error leyendo columna de línea '{col_linea}': {e}")

                if lineas_detectadas:
                    # Actualizar la configuración con las líneas detectadas
                    config_actual = self.config_service.get_config()
                    lineas_config = dict(config_actual.get("lineas", {}))
                    tipos_config = config_actual.get("tipos_bus", {})
                    tipos_por_defecto = list(tipos_config.keys()) or TIPOS_BUS_DISPONIBLES

                    for linea in lineas_detectadas:
                        if linea not in lineas_config:
                            lineas_config[linea] = {
                                "tipos_permitidos": tipos_por_defecto,
                            }

                    # Guardar en la configuración interna; el persistido se hace fuera
                    self.config_service._config_data["lineas"] = lineas_config  # type: ignore[attr-defined]

                    logger.info(
                        f"Líneas detectadas desde Excel: {sorted(lineas_detectadas)} "
                        f"(total ahora en config: {len(lineas_config)})"
                    )
            else:
                logger.warning(
                    "El Excel de viajes no tiene columna 'Linea'; "
                    "no se pueden inferir las líneas habilitadas automáticamente."
                )
            
            logger.info(
                f"Depósitos activos (no se modifican desde Excel de salidas): "
                f"{self.config_service.get_active_deposits()}"
            )
            logger.info(f"Nodos detectados: {nodos_detectados}")
            
            return sorted(nodos_detectados)
            
        except FileNotFoundError as e:
            logger.error(f"Archivo no encontrado: {e}")
            return []
        except pd.errors.EmptyDataError:
            logger.error("El archivo Excel está vacío o no tiene datos")
            return []
        except pd.errors.ParserError as e:
            logger.error(f"Error al parsear el archivo Excel: {e}")
            return []
        except KeyError as e:
            logger.error(f"Error: columna no encontrada en el archivo Excel: {e}")
            logger.info("Asegúrese de que el archivo tenga columnas 'Origen' y 'Destino'")
            return []
        except Exception as e:
            logger.error(f"Error detectando nodos y líneas desde Excel: {e}")
            logger.error(traceback.format_exc())
            return []
    
    def get_trips_file_path(self) -> Optional[str]:
        """
        Obtiene la ruta del archivo de viajes: siempre datos_salidas.xlsx (sin copias ni puntero).
        
        Returns:
            Ruta absoluta del archivo si existe, None en caso contrario
        """
        if os.path.exists(self._trips_file_path):
            return os.path.abspath(self._trips_file_path)
        return None
    
    def generate_trips_template(self) -> io.BytesIO:
        """
        Genera un template Excel para cargar viajes comerciales.
        
        Returns:
            BytesIO con el archivo Excel generado
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "ViajesComerciales"
        
        # Encabezados
        headers = ["Linea", "Sentido", "Origen", "Destino", "Hora Inicio", "Hora Fin", "Kilometros"]
        ws.append(headers)
        
        # Formatear encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Agregar filas de ejemplo
        ejemplos = [
            ["712", "1", "TILVOL", "ROTPIR", "06:00", "07:25", "42.34"],
            ["712", "2", "ROTPIR", "TILVOL", "07:30", "08:55", "41.532"],
            ["712", "1", "TILVOL", "ROTPIR", "09:00", "10:25", "42.34"],
        ]
        
        for ejemplo in ejemplos:
            ws.append(ejemplo)
        
        # Ajustar ancho de columnas
        column_widths = {
            "A": 12, "B": 12, "C": 15, "D": 15,
            "E": 15, "F": 15, "G": 15,
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Agregar hoja de instrucciones
        ws_instrucciones = wb.create_sheet("Instrucciones")
        instrucciones = [
            ["INSTRUCCIONES PARA COMPLETAR EL TEMPLATE"],
            [""],
            ["Columnas requeridas:"],
            ["- Linea: Número o código de la línea de transporte"],
            ["- Sentido: 1 o 2 (sentido de la ruta)"],
            ["- Origen: Nombre del nodo de origen"],
            ["- Destino: Nombre del nodo de destino"],
            ["- Hora Inicio: Hora de inicio del viaje (formato HH:MM, ej: 06:00)"],
            ["- Hora Fin: Hora de fin del viaje (formato HH:MM, ej: 07:25)"],
            ["- Kilometros: Distancia del viaje en kilómetros (puede usar decimales)"],
            [""],
            ["Notas:"],
            ["- Puede eliminar las filas de ejemplo y agregar sus propios viajes"],
            ["- Asegúrese de que los nodos (Origen/Destino) estén configurados en la pestaña 'Configuración'"],
            ["- El formato de hora debe ser HH:MM (24 horas)"],
            ["- Los kilómetros pueden tener decimales (use punto como separador)"],
        ]
        
        for fila in instrucciones:
            ws_instrucciones.append(fila)
        
        ws_instrucciones["A1"].font = Font(bold=True, size=14)
        ws_instrucciones.column_dimensions["A"].width = 80
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output


class ExcelConfigService:
    """
    Servicio para generar y procesar archivos Excel de configuración.
    """
    
    def __init__(self, config_service: ConfigService, excel_trips_service: Optional["ExcelTripsService"] = None):
        """
        Inicializa el servicio de Excel de configuración.
        
        Args:
            config_service: Instancia de ConfigService
            excel_trips_service: Opcional. Si se proporciona, al importar config se fusionan
                nodos detectados desde datos_salidas.xlsx (precarga).
        """
        self.config_service = config_service
        self.excel_trips_service = excel_trips_service
    
    def generate_config_template(self) -> io.BytesIO:
        """
        Genera un template Excel con la configuración actual.
        
        Returns:
            BytesIO con el archivo Excel generado
        """
        config = self.config_service.get_config()
        wb = Workbook()
        
        # Hoja 1: Configuración General
        ws1 = wb.active
        ws1.title = "ConfiguracionGeneral"
        ws1.append(["Parámetro", "Valor"])
        ws1.append(["deposito", config.get("deposito", "Deposito Pie Andino")])
        ws1.append(["limite_jornada", config.get("limite_jornada", 720)])
        ws1.append(["tiempo_toma", config.get("tiempo_toma", 15)])
        ws1.append(["max_inicio_jornada_conductor", config.get("max_inicio_jornada_conductor", "23:59")])
        ws1.append(["interlineado_global", "SI" if config.get("interlineado_global", False) else "NO"])
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws1.column_dimensions["A"].width = 20
        ws1.column_dimensions["B"].width = 30
        
        # Hoja 2: Nodos
        ws2 = wb.create_sheet("Nodos")
        ws2.append(["Nodo"])
        for nodo in config.get("nodos", []):
            ws2.append([nodo])
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions["A"].width = 30
        
        # Hoja 3: Vacíos
        ws3 = wb.create_sheet("Vacios")
        ws3.append(["Origen", "Destino", "Habilitado", "Inicio", "Fin", "Tiempo", "Km"])
        vacios = config.get("vacios", {})
        for conexion, datos in sorted(vacios.items()):
            origen, destino = conexion.split("_", 1) if "_" in conexion else (conexion, "")
            habilitado = datos.get("habilitado", True)
            franjas = datos.get("franjas", [])
            if not franjas:
                ws3.append([origen, destino, "SI" if habilitado else "NO", "", "", "", ""])
            else:
                for franja in franjas:
                    ws3.append([
                        origen, destino,
                        "SI" if habilitado else "NO",
                        franja.get("inicio", ""),
                        franja.get("fin", ""),
                        franja.get("tiempo", ""),
                        franja.get("km", "")
                    ])
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws3.column_dimensions["A"].width = 25
        ws3.column_dimensions["B"].width = 25
        ws3.column_dimensions["C"].width = 12
        for col in ["D", "E", "F", "G"]:
            ws3.column_dimensions[col].width = 15
        
        # Hoja 4: Desplazamientos
        ws4 = wb.create_sheet("Desplazamientos")
        ws4.append(["Origen", "Destino", "Habilitado", "Tiempo"])
        desplazamientos = config.get("desplazamientos", {})
        for conexion, datos in sorted(desplazamientos.items()):
            origen, destino = conexion.split("_", 1) if "_" in conexion else (conexion, "")
            ws4.append([
                origen, destino,
                "SI" if datos.get("habilitado", True) else "NO",
                datos.get("tiempo", 30)
            ])
        for cell in ws4[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws4.column_dimensions["A"].width = 25
        ws4.column_dimensions["B"].width = 25
        ws4.column_dimensions["C"].width = 12
        ws4.column_dimensions["D"].width = 15
        
        # Hoja 5b: Puntos de Relevo
        ws5b = wb.create_sheet("PuntosRelevo")
        ws5b.append(["Nodo", "Es Punto de Relevo"])
        depositos_nombres = set()
        for dep in (config.get("depositos") or []):
            if isinstance(dep, dict) and dep.get("nombre"):
                depositos_nombres.add(str(dep["nombre"]).strip())
        if config.get("deposito"):
            depositos_nombres.add(str(config["deposito"]).strip())
        puntos_relevo_set = set(str(p).strip() for p in (config.get("puntos_relevo") or []))
        for nodo in config.get("nodos", []):
            if nodo and str(nodo).strip() not in depositos_nombres:
                ws5b.append([nodo, "SI" if str(nodo).strip() in puntos_relevo_set else "NO"])
        for cell in ws5b[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws5b.column_dimensions["A"].width = 30
        ws5b.column_dimensions["B"].width = 20
        
        # Hoja 5: Paradas
        ws5 = wb.create_sheet("Paradas")
        ws5.append(["Nodo", "Min", "Max"])
        paradas = config.get("paradas", {})
        for nodo, datos in sorted(paradas.items()):
            ws5.append([nodo, datos.get("min", 5), datos.get("max", 120)])
        for cell in ws5[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws5.column_dimensions["A"].width = 30
        ws5.column_dimensions["B"].width = 12
        ws5.column_dimensions["C"].width = 12
        
        # Hoja 6: Depósitos
        ws6 = wb.create_sheet("Depositos")
        encabezados_depositos = ["Nombre", "Max Buses", "Permite Recarga", "Posiciones Recarga"]
        for tipo in TIPOS_BUS_DISPONIBLES:
            encabezados_depositos.append(f"Flota {tipo}")
        ws6.append(encabezados_depositos)
        
        depositos = config.get("depositos") or []
        if not depositos and config.get("deposito"):
            depositos = [{
                "nombre": config.get("deposito"),
                "max_buses": 200,
                "permite_recarga": True,
                "posiciones_recarga": 0,
                "flota_por_tipo": config.get("flota_por_tipo", {}),
            }]
        
        for dep in depositos:
            fila = [
                dep.get("nombre", ""),
                dep.get("max_buses", ""),
                "SI" if dep.get("permite_recarga", True) else "NO",
                dep.get("posiciones_recarga", 0),
            ]
            flota_dep = dep.get("flota_por_tipo", {})
            for tipo in TIPOS_BUS_DISPONIBLES:
                fila.append(flota_dep.get(tipo, 0))
            ws6.append(fila)
        
        for cell in ws6[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, len(encabezados_depositos) + 1):
            ws6.column_dimensions[chr(64 + col)].width = 18 if col <= 4 else 14
        
        # Hoja 7: Tipos de Bus
        ws7 = wb.create_sheet("TiposBus")
        ws7.append([
            "Tipo", "Habilitado", "Descripción", "Autonomía (km)", "Capacidad",
            "Es Eléctrico", "Carga Inicial (%)", "Consumo (% por km)",
            "Límite Operación (%)", "Min Entrada Recarga (%)",
            "Max Entrada Recarga (%)", "Tasa Recarga (%/min)",
            "Ventana Inicio", "Ventana Fin",
        ])
        
        tipos_config = config.get("tipos_bus", {})
        for tipo in TIPOS_BUS_DISPONIBLES:
            datos = tipos_config.get(tipo, {})
            habilitado = "SI" if tipo in tipos_config else "NO"
            descripcion = datos.get("descripcion", "")
            autonomia = datos.get("autonomia_km", "")
            capacidad = datos.get("capacidad_pasajeros", "")
            es_electrico = bool(datos.get("es_electrico", tipo == "BE"))
            parametros = datos.get("parametros_electricos", {}) if es_electrico else {}
            
            ws7.append([
                tipo, habilitado, descripcion, autonomia, capacidad,
                "SI" if es_electrico else "NO",
                parametros.get("carga_inicial", ""),
                parametros.get("consumo_por_km", ""),
                parametros.get("limite_operacion", ""),
                parametros.get("min_entrada_recarga", ""),
                parametros.get("max_entrada_recarga", ""),
                parametros.get("tasa_recarga_por_minuto", ""),
                (parametros.get("ventana_recarga") or {}).get("inicio", ""),
                (parametros.get("ventana_recarga") or {}).get("fin", ""),
            ])
        
        for cell in ws7[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws7.column_dimensions["A"].width = 10
        ws7.column_dimensions["B"].width = 12
        ws7.column_dimensions["C"].width = 35
        for col in ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
            ws7.column_dimensions[col].width = 18
        
        # Hoja 8: Líneas
        ws8 = wb.create_sheet("Lineas")
        ws8.append(["Linea", "Tipos Permitidos (coma)"])
        lineas_config = config.get("lineas", {})
        for linea, datos in sorted(lineas_config.items()):
            ws8.append([linea, ", ".join(datos.get("tipos_permitidos", []))])
        for cell in ws8[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws8.column_dimensions["A"].width = 18
        ws8.column_dimensions["B"].width = 45

        # Hoja 8b: Grupos de Líneas
        ws8b = wb.create_sheet("GruposLineas")
        ws8b.append(["Grupo", "Linea"])
        grupos_lineas = config.get("grupos_lineas", {}) or {}
        for nombre_grupo, lineas in sorted(grupos_lineas.items()):
            for linea in (lineas or []):
                ws8b.append([nombre_grupo, linea])
        for cell in ws8b[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws8b.column_dimensions["A"].width = 26
        ws8b.column_dimensions["B"].width = 22

        # Hoja 8d: Límites de jornada por grupo de línea
        ws8d = wb.create_sheet("JornadaPorGrupo")
        ws8d.append(["Grupo", "LimiteJornadaMin"])
        limites_grupo = config.get("limite_jornada_por_grupo_linea", {}) or {}
        limite_global = int(config.get("limite_jornada", 720) or 720)
        for nombre_grupo, lineas in sorted(grupos_lineas.items()):
            ws8d.append([nombre_grupo, int(limites_grupo.get(nombre_grupo, limite_global))])
        for cell in ws8d[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws8d.column_dimensions["A"].width = 26
        ws8d.column_dimensions["B"].width = 22

        # Hoja 8c: Fase 3 Unión de Conductores
        ws8c = wb.create_sheet("Fase3Union")
        ws8c.append(["Parametro", "Valor"])
        f3_cfg = config.get("fase_3_union_conductores", {}) or {}
        ws8c.append(["union_solo_por_deposito", "SI" if f3_cfg.get("union_solo_por_deposito", False) else "NO"])
        ws8c.append(["parada_larga_umbral_union", f3_cfg.get("parada_larga_umbral_union", 60)])
        ws8c.append(["parada_larga_excepcion_depot_min", f3_cfg.get("parada_larga_excepcion_depot_min", 120)])
        ws8c.append(["max_rondas_union", f3_cfg.get("max_rondas_union", 1500)])
        ws8c.append(["timeout_ortools_segundos", f3_cfg.get("timeout_ortools_segundos", 180)])
        for cell in ws8c[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws8c.column_dimensions["A"].width = 36
        ws8c.column_dimensions["B"].width = 24

        # Hoja 9: Tipos de Conductor
        ws9 = wb.create_sheet("TiposConductor")
        ws9.append(["ID", "Nombre", "Ingreso Min (HH:MM)", "Ingreso Max (HH:MM)", "Fin Jornada Min (HH:MM)", "Fin Jornada Max (HH:MM)"])
        tipos_conductor = config.get("tipos_conductor") or []
        for tc in tipos_conductor:
            ri = tc.get("rango_ingreso") or {}
            rf = tc.get("rango_fin_jornada") or {}
            ws9.append([
                tc.get("id", ""),
                tc.get("nombre", ""),
                ri.get("min", "00:00"),
                ri.get("max", "23:59"),
                rf.get("min", "00:00"),
                rf.get("max", "23:59"),
            ])
        for cell in ws9[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws9.column_dimensions[col].width = 22

        # Hoja 10: Guía detallada de uso (qué es, para qué sirve y efecto de cambios)
        ws10 = wb.create_sheet("Guia")
        ws10.append(["Solapa", "Campo/Sección", "Qué es", "Para qué sirve", "Qué pasa si se modifica"])
        guia_filas = [
            ["ConfiguracionGeneral", "deposito", "Depósito base por defecto", "Punto de referencia operativo", "Cambia el origen/final esperado de jornadas y conexiones."],
            ["ConfiguracionGeneral", "limite_jornada", "Tope de minutos por jornada de conductor", "Cumplir regla laboral/operativa", "Subirlo reduce conductores pero alarga jornadas; bajarlo aumenta cortes y conductores."],
            ["ConfiguracionGeneral", "tiempo_toma", "Minutos de preparación antes de operar", "Modelar inicio real de jornada", "Más alto exige iniciar antes; más bajo da más holgura temporal."],
            ["Depositos", "max_buses", "Límite de buses por depósito", "Restringir recursos de flota por base", "Si no alcanza para cubrir viajes, Fase 1 falla en modo estricto."],
            ["ConfiguracionGeneral", "max_inicio_jornada_conductor", "Hora máxima permitida para iniciar jornada", "Controlar entradas tardías", "Más temprano restringe planificación; más tarde la flexibiliza."],
            ["ConfiguracionGeneral", "interlineado_global", "Permite mezclar líneas libremente", "Control de compatibilidad entre líneas", "En SI ignora grupos; en NO usa GruposLineas."],
            ["Nodos", "Nodo", "Catálogo de puntos operativos", "Base para vacíos, paradas, desplazamientos", "Nodos faltantes provocan configuraciones incompletas o imposibles."],
            ["Vacios", "Origen/Destino/Habilitado/Franja/Tiempo/Km", "Traslado sin pasajeros entre nodos", "Conectar viajes y retornos", "Tiempos altos dificultan conexiones; valores irreales degradan factibilidad."],
            ["Desplazamientos", "Origen/Destino/Habilitado/Tiempo", "Movimiento de conductor sin bus", "Cerrar encadenamientos de turnos", "Tiempos altos reducen uniones y aumentan conductores."],
            ["PuntosRelevo", "Nodo + Es Punto de Relevo", "Nodos habilitados para intercambio", "Permitir relevos operativos", "Más puntos facilita cobertura; menos puntos restringe combinaciones."],
            ["Paradas", "Min/Max por nodo", "Ventana de espera permitida", "Respetar operación por terminal/nodo", "Max bajo puede aumentar buses; max alto aumenta flexibilidad."],
            ["Depositos", "Capacidad, recarga y flota por tipo", "Restricciones físicas de base", "Controlar factibilidad real de recursos", "Capacidad baja puede volver problema infactible."],
            ["TiposBus", "Parámetros por tipo (incluye eléctricos)", "Definición técnica de cada flota", "Calcular consumo, recarga y elegibilidad", "Parámetros eléctricos impactan autonomía y necesidad de recarga."],
            ["Lineas", "Tipos Permitidos", "Compatibilidad línea-tipo de bus", "Evitar asignaciones no válidas", "Reglas más estrictas reducen opciones de asignación."],
            ["GruposLineas", "Grupo/Linea", "Agrupación para interlineado", "Permitir mezcla solo dentro de grupo", "Grupos más cerrados restringen reutilización y unión."],
            ["JornadaPorGrupo", "Grupo/LimiteJornadaMin", "Límite de jornada específico por grupo", "Aplicar reglas laborales distintas por operación", "Si no existe un grupo aquí, usa limite_jornada global."],
            ["Fase3Union", "union_solo_por_deposito", "Unión de turnos solo vía depósito", "Aumentar robustez operacional de uniones", "En SI reduce uniones; en NO permite conexiones directas."],
            ["Fase3Union", "parada_larga_umbral_union", "Umbral de espera para unir turnos", "Evitar uniones con inactividad excesiva", "Más alto permite más uniones; más bajo las restringe."],
            ["Fase3Union", "parada_larga_excepcion_depot_min", "Excepción en depósito para parada larga", "Flexibilizar unión cuando espera es en depósito", "Más alto permite más casos especiales en depósito."],
            ["Fase3Union", "max_rondas_union", "Iteraciones máximas de mejora", "Profundizar búsqueda de uniones", "Más rondas puede mejorar resultado pero demora más."],
            ["Fase3Union", "timeout_ortools_segundos", "Tiempo máximo del solver", "Limitar cómputo en optimización", "Más segundos puede mejorar calidad, con mayor tiempo de ejecución."],
            ["TiposConductor", "Rangos de ingreso/fin por perfil", "Reglas por tipo de conductor", "Respetar ventanas de operación", "Ventanas más estrictas pueden elevar cantidad de conductores necesarios."],
        ]
        for fila in guia_filas:
            ws10.append(fila)
        for cell in ws10[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws10.column_dimensions["A"].width = 24
        ws10.column_dimensions["B"].width = 34
        ws10.column_dimensions["C"].width = 34
        ws10.column_dimensions["D"].width = 34
        ws10.column_dimensions["E"].width = 58
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def process_config_upload(self, archivo: Optional[FileStorage]) -> Dict[str, Any]:
        """
        Procesa un archivo Excel de configuración y actualiza la configuración.
        
        Args:
            archivo: Archivo Excel de configuración subido
            
        Returns:
            Diccionario con resultado de la operación
        """
        es_valido, mensaje_error = validate_excel_file(archivo)
        if not es_valido:
            return {"success": False, "message": mensaje_error}
        
        try:
            # Guardar temporalmente
            archivo_path = "configuracion_temp.xlsx"
            archivo.save(archivo_path)
            
            # Procesar Excel
            nueva_config = self._read_config_from_excel(archivo_path)
            
            if not nueva_config:
                return {"success": False, "message": "Error al procesar el archivo Excel de configuración"}
            
            # Validar estructura básica
            from ..utils.validators import validate_config_data
            es_valido, mensaje_error = validate_config_data(nueva_config)
            if not es_valido:
                return {"success": False, "message": mensaje_error}
            
            # PRECARGA: Fusionar nodos desde datos_salidas.xlsx si existe
            if self.excel_trips_service:
                trips_path = self.excel_trips_service.get_trips_file_path()
                if trips_path and os.path.exists(trips_path):
                    nodos_salidas = self.excel_trips_service._detect_nodes_from_excel(trips_path)
                    if nodos_salidas:
                        nodos_actuales = set(nueva_config.get("nodos", []))
                        nodos_fusionados = sorted(nodos_actuales | set(nodos_salidas))
                        nueva_config["nodos"] = nodos_fusionados
                        logger.info(
                            f"Precarga desde datos_salidas: {len(nodos_salidas)} nodos detectados, "
                            f"total fusionado: {len(nodos_fusionados)}"
                        )
            
            # Actualizar configuración
            self.config_service._config_data = nueva_config
            self.config_service.regenerate_connections()
            
            if self.config_service.save_config():
                # Limpiar archivo temporal
                try:
                    os.remove(archivo_path)
                except:
                    pass
                
                return {
                    "success": True,
                    "message": f"Configuración importada exitosamente. {len(nueva_config.get('nodos', []))} nodos configurados."
                }
            else:
                return {"success": False, "message": "Error al guardar la configuración importada"}
                
        except Exception as e:
            logger.error(f"Error importando configuración: {e}")
            traceback.print_exc()
            # Limpiar archivo temporal si existe
            try:
                if os.path.exists("configuracion_temp.xlsx"):
                    os.remove("configuracion_temp.xlsx")
            except:
                pass
            return {"success": False, "message": f"Error importando configuración: {str(e)}"}
    
    def _read_config_from_excel(self, archivo_path: str) -> Optional[Dict[str, Any]]:
        """
        Lee configuración desde un archivo Excel.
        
        Args:
            archivo_path: Ruta al archivo Excel
            
        Returns:
            Diccionario de configuración o None si hay error
        """
        try:
            wb = load_workbook(archivo_path, data_only=True)
            base_cfg = self.config_service.get_config() or {}
            nueva_config = copy.deepcopy(base_cfg)
            nueva_config.setdefault("deposito", "Deposito Pie Andino")
            nueva_config.setdefault("limite_jornada", 720)
            nueva_config.setdefault("tiempo_toma", 15)
            nueva_config.setdefault("max_inicio_jornada_conductor", "23:59")
            nueva_config.setdefault("nodos", [])
            nueva_config.setdefault("vacios", {})
            nueva_config.setdefault("desplazamientos", {})
            nueva_config.setdefault("paradas", {})
            nueva_config.setdefault("puntos_relevo", [])
            nueva_config.setdefault("tipos_bus", {})
            nueva_config.setdefault("lineas", {})
            nueva_config.setdefault("depositos", [])
            nueva_config.setdefault("flota_por_tipo", {})
            nueva_config.setdefault("tipos_conductor", [])
            nueva_config.setdefault("interlineado_global", False)
            nueva_config.setdefault("grupos_lineas", {})
            nueva_config.setdefault("limite_jornada_por_grupo_linea", {})
            nueva_config.setdefault("fase_3_union_conductores", {})
            nodos_recolectados: set = set()
            
            # Leer Configuración General
            if "ConfiguracionGeneral" in wb.sheetnames:
                try:
                    ws = wb["ConfiguracionGeneral"]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or len(row) < 2:
                            continue
                        if row[0] and row[1] is not None:
                            try:
                                param = str(row[0]).strip().lower()
                                valor = row[1]
                                if param == "deposito":
                                    dep_name = str(valor).strip() if valor else "Deposito Pie Andino"
                                    nueva_config["deposito"] = dep_name
                                    nodos_recolectados.add(dep_name)
                                elif param == "limite_jornada":
                                    nueva_config["limite_jornada"] = int(float(valor)) if valor else 720
                                elif param == "tiempo_toma":
                                    nueva_config["tiempo_toma"] = int(float(valor)) if valor else 15
                                elif param == "max_inicio_jornada_conductor":
                                    nueva_config["max_inicio_jornada_conductor"] = str(valor).strip() if valor else "23:59"
                                elif param == "interlineado_global":
                                    nueva_config["interlineado_global"] = _convertir_habilitado_a_booleano(valor, valor_por_defecto=False)
                            except (ValueError, TypeError) as e:
                                logger.warning(f"Error procesando parámetro {row[0]}: {e}")
                                continue
                except Exception as e:
                    logger.warning(f"Error leyendo ConfiguracionGeneral: {e}")
            
            # Leer Nodos (hoja explícita) y recolectar de Vacíos, Desplazamientos, Paradas, Depósitos
            if "Nodos" in wb.sheetnames:
                ws = wb["Nodos"]
                max_row = getattr(ws, "max_row", None) or 10000
                for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
                    if row and row[0] and str(row[0]).strip():
                        nodos_recolectados.add(str(row[0]).strip())
            
            # Leer Vacíos (simplificado - solo formato nuevo)
            if "Vacios" in wb.sheetnames:
                ws = wb["Vacios"]
                conexiones_dict = {}
                primera_fila = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                formato_nuevo = primera_fila and len(primera_fila) > 1 and str(primera_fila[0]).strip().lower() == "origen"
                max_row = getattr(ws, "max_row", None) or 10000
                
                for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
                    if formato_nuevo:
                        origen = str(row[0]).strip() if row[0] else None
                        destino = str(row[1]).strip() if row[1] else None
                        # Leer valor de habilitado usando función auxiliar
                        valor_habilitado = row[2] if len(row) > 2 else None
                        habilitado = _convertir_habilitado_a_booleano(valor_habilitado, valor_por_defecto=True)
                        
                        inicio = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                        fin = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                        tiempo = row[5] if len(row) > 5 and row[5] is not None else None
                        km = row[6] if len(row) > 6 and row[6] is not None else None
                    else:
                        conexion_str = str(row[0]).strip() if row[0] else None
                        if conexion_str and "_" in conexion_str:
                            origen, destino = conexion_str.split("_", 1)
                        else:
                            origen = conexion_str
                            destino = None
                        
                        # Leer valor de habilitado usando función auxiliar
                        valor_habilitado = row[1] if len(row) > 1 else None
                        habilitado = _convertir_habilitado_a_booleano(valor_habilitado, valor_por_defecto=True)
                        
                        inicio = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                        fin = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                        tiempo = row[4] if len(row) > 4 and row[4] is not None else None
                        km = row[5] if len(row) > 5 and row[5] is not None else None
                    
                    if origen and destino:
                        nodos_recolectados.add(origen)
                        nodos_recolectados.add(destino)
                        conexion = f"{origen}_{destino}"
                        
                        if conexion not in conexiones_dict:
                            conexiones_dict[conexion] = {"habilitado": habilitado, "franjas": []}
                        else:
                            if len(conexiones_dict[conexion]["franjas"]) == 0:
                                conexiones_dict[conexion]["habilitado"] = habilitado
                        
                        if inicio and fin:
                            try:
                                tiempo_val = float(tiempo) if tiempo is not None else 30.0
                            except (TypeError, ValueError):
                                tiempo_val = 30.0
                            try:
                                km_val = float(km) if km is not None else 0.0
                            except (TypeError, ValueError):
                                km_val = 0.0
                            conexiones_dict[conexion]["franjas"].append({
                                "inicio": inicio,
                                "fin": fin,
                                "tiempo": tiempo_val,
                                "km": km_val
                            })
                
                nueva_config["vacios"] = conexiones_dict
            
            # Leer Desplazamientos
            if "Desplazamientos" in wb.sheetnames:
                ws = wb["Desplazamientos"]
                max_row = getattr(ws, "max_row", None) or 10000
                for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
                    if not row or not row[0] or not row[1]:
                        continue
                    origen = str(row[0]).strip()
                    destino = str(row[1]).strip()
                    if origen and destino:
                        nodos_recolectados.add(origen)
                        nodos_recolectados.add(destino)
                        conexion = f"{origen}_{destino}"
                        valor_habilitado = row[2] if len(row) > 2 else None
                        habilitado = _convertir_habilitado_a_booleano(valor_habilitado, valor_por_defecto=True)
                        tiempo = float(row[3]) if len(row) > 3 and row[3] is not None else 30.0
                        nueva_config["desplazamientos"][conexion] = {
                            "habilitado": habilitado,
                            "tiempo": tiempo
                        }
            
            # Leer Paradas
            if "Paradas" in wb.sheetnames:
                ws = wb["Paradas"]
                max_row = getattr(ws, "max_row", None) or 10000
                for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
                    if row and row[0]:
                        nodo = str(row[0]).strip()
                        if nodo:
                            nodos_recolectados.add(nodo)
                            min_val = int(float(row[1])) if len(row) > 1 and row[1] is not None else 5
                            max_val = int(float(row[2])) if len(row) > 2 and row[2] is not None else 120
                            nueva_config["paradas"][nodo] = {"min": min_val, "max": max_val}
            
            # Leer Puntos de Relevo
            if "PuntosRelevo" in wb.sheetnames:
                ws = wb["PuntosRelevo"]
                puntos_relevo = []
                max_row = getattr(ws, "max_row", None) or 10000
                for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
                    if row and row[0]:
                        nodo = str(row[0]).strip()
                        if nodo:
                            nodos_recolectados.add(nodo)
                            es_relevo = _convertir_habilitado_a_booleano(row[1] if len(row) > 1 else None, valor_por_defecto=False)
                            if es_relevo:
                                puntos_relevo.append(nodo)
                nueva_config["puntos_relevo"] = puntos_relevo
            
            # Leer Depósitos (simplificado)
            if "Depositos" in wb.sheetnames:
                try:
                    ws = wb["Depositos"]
                    depositos = []
                    flota_total = {tipo: 0 for tipo in TIPOS_BUS_DISPONIBLES}
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or not row[0]:
                            continue
                        try:
                            nombre = str(row[0]).strip()
                            if not nombre:
                                continue
                            nodos_recolectados.add(nombre)
                            max_buses = int(float(row[1])) if row[1] is not None else 0
                            permite = str(row[2]).strip().upper() == "SI" if row[2] else True
                            posiciones = int(float(row[3])) if row[3] is not None else 0
                            flota_por_tipo = {}
                            for idx, tipo in enumerate(TIPOS_BUS_DISPONIBLES):
                                valor = row[4 + idx] if len(row) > 4 + idx else 0
                                cantidad = max(0, int(float(valor)) if valor else 0)
                                flota_por_tipo[tipo] = cantidad
                                flota_total[tipo] = flota_total.get(tipo, 0) + cantidad
                            depositos.append({
                                "nombre": nombre,
                                "max_buses": max_buses,
                                "permite_recarga": permite,
                                "posiciones_recarga": posiciones,
                                "flota_por_tipo": flota_por_tipo,
                            })
                        except (ValueError, TypeError, IndexError) as e:
                            logger.warning(f"Error procesando fila de depósito: {e}")
                            continue
                    if depositos:
                        nueva_config["depositos"] = depositos
                        nueva_config["flota_por_tipo"] = flota_total
                        # Si hay depósitos, usar el primero como depósito principal
                        if depositos and not nueva_config.get("deposito") or nueva_config.get("deposito") == "Deposito Pie Andino":
                            nueva_config["deposito"] = depositos[0]["nombre"]
                except Exception as e:
                    logger.warning(f"Error leyendo Depósitos: {e}")
            
            # Leer Tipos de Bus (simplificado)
            if "TiposBus" in wb.sheetnames:
                ws = wb["TiposBus"]
                tipos_result = {}
                max_row = getattr(ws, "max_row", None) or 100
                for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
                    if not row or not row[0]:
                        continue
                    tipo = str(row[0]).strip().upper()
                    if not tipo:
                        continue
                    habilitado = str(row[1]).strip().upper() != "NO" if row[1] else True
                    if not habilitado:
                        continue
                    entry = {}
                    if row[2]:
                        entry["descripcion"] = str(row[2]).strip()
                    autonomia = float(row[3]) if row[3] else None
                    if autonomia is not None:
                        entry["autonomia_km"] = autonomia
                    capacidad = int(row[4]) if row[4] else 0
                    if capacidad:
                        entry["capacidad_pasajeros"] = capacidad
                    es_electrico = str(row[5]).strip().upper() == "SI" if row[5] else (tipo == "BE")
                    entry["es_electrico"] = es_electrico
                    if es_electrico:
                        parametros = {
                            "carga_inicial": float(row[6]) if row[6] else 95,
                            "consumo_por_km": float(row[7]) if row[7] else 0.5,
                            "limite_operacion": float(row[8]) if row[8] else 30,
                            "min_entrada_recarga": float(row[9]) if row[9] else 60,
                            "max_entrada_recarga": float(row[10]) if row[10] else 80,
                            "tasa_recarga_por_minuto": float(row[11]) if row[11] else 1.25,
                            "ventana_recarga": {
                                "inicio": str(row[12]).strip() if row[12] else "09:00",
                                "fin": str(row[13]).strip() if row[13] else "18:00",
                            },
                        }
                        entry["parametros_electricos"] = parametros
                    tipos_result[tipo] = entry
                nueva_config["tipos_bus"] = tipos_result
            
            # Leer Líneas
            if "Lineas" in wb.sheetnames:
                ws = wb["Lineas"]
                lineas_result = {}
                max_row = getattr(ws, "max_row", None) or 500
                for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
                    if not row or not row[0]:
                        continue
                    linea = str(row[0]).strip()
                    if not linea:
                        continue
                    tipos = []
                    if len(row) > 1 and row[1]:
                        tipos = [t.strip().upper() for t in str(row[1]).split(",") if t and str(t).strip()]
                    if not tipos:
                        tipos = list(nueva_config.get("tipos_bus", {}).keys()) or TIPOS_BUS_DISPONIBLES
                    lineas_result[linea] = {"tipos_permitidos": tipos}
                nueva_config["lineas"] = lineas_result

            # Leer Grupos de Líneas
            if "GruposLineas" in wb.sheetnames:
                ws = wb["GruposLineas"]
                grupos: Dict[str, List[str]] = {}
                max_row = getattr(ws, "max_row", None) or 500
                for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
                    if not row:
                        continue
                    grupo = str(row[0]).strip() if len(row) > 0 and row[0] else ""
                    linea = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    if not grupo or not linea:
                        continue
                    grupos.setdefault(grupo, [])
                    if linea not in grupos[grupo]:
                        grupos[grupo].append(linea)
                nueva_config["grupos_lineas"] = {g: sorted(ls) for g, ls in grupos.items()}

            # Leer límite jornada por grupo de línea
            if "JornadaPorGrupo" in wb.sheetnames:
                ws = wb["JornadaPorGrupo"]
                limites_grupo: Dict[str, int] = {}
                max_row = getattr(ws, "max_row", None) or 500
                for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
                    if not row:
                        continue
                    grupo = str(row[0]).strip() if len(row) > 0 and row[0] else ""
                    if not grupo:
                        continue
                    try:
                        limite = int(float(row[1])) if len(row) > 1 and row[1] is not None else int(nueva_config.get("limite_jornada", 720))
                    except Exception:
                        limite = int(nueva_config.get("limite_jornada", 720))
                    limites_grupo[grupo] = limite
                nueva_config["limite_jornada_por_grupo_linea"] = limites_grupo

            # Leer Configuración Fase 3 Unión
            if "Fase3Union" in wb.sheetnames:
                ws = wb["Fase3Union"]
                f3_cfg = dict(nueva_config.get("fase_3_union_conductores") or {})
                max_row = getattr(ws, "max_row", None) or 200
                for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
                    if not row or not row[0]:
                        continue
                    param = str(row[0]).strip().lower()
                    valor = row[1] if len(row) > 1 else None
                    if param == "union_solo_por_deposito":
                        f3_cfg["union_solo_por_deposito"] = _convertir_habilitado_a_booleano(valor, valor_por_defecto=False)
                    elif param == "parada_larga_umbral_union":
                        try:
                            f3_cfg["parada_larga_umbral_union"] = int(float(valor))
                        except (TypeError, ValueError):
                            pass
                    elif param == "parada_larga_excepcion_depot_min":
                        try:
                            f3_cfg["parada_larga_excepcion_depot_min"] = int(float(valor))
                        except (TypeError, ValueError):
                            pass
                    elif param == "max_rondas_union":
                        try:
                            f3_cfg["max_rondas_union"] = int(float(valor))
                        except (TypeError, ValueError):
                            pass
                    elif param == "timeout_ortools_segundos":
                        try:
                            f3_cfg["timeout_ortools_segundos"] = int(float(valor))
                        except (TypeError, ValueError):
                            pass
                nueva_config["fase_3_union_conductores"] = f3_cfg

            # Leer Tipos de Conductor
            if "TiposConductor" in wb.sheetnames:
                ws = wb["TiposConductor"]
                tipos_conductor = []
                max_row = getattr(ws, "max_row", None) or 50
                for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
                    if not row or (row[0] is None and row[1] is None):
                        continue
                    tipo_id = str(row[0]).strip() if row[0] else ""
                    nombre = str(row[1]).strip() if len(row) > 1 and row[1] else tipo_id
                    if not tipo_id and not nombre:
                        continue
                    if not tipo_id:
                        tipo_id = nombre
                    ing_min = str(row[2]).strip() if len(row) > 2 and row[2] else "00:00"
                    ing_max = str(row[3]).strip() if len(row) > 3 and row[3] else "23:59"
                    fin_min = str(row[4]).strip() if len(row) > 4 and row[4] else "00:00"
                    fin_max = str(row[5]).strip() if len(row) > 5 and row[5] else "23:59"
                    tipos_conductor.append({
                        "id": tipo_id,
                        "nombre": nombre,
                        "rango_ingreso": {"min": ing_min, "max": ing_max},
                        "rango_fin_jornada": {"min": fin_min, "max": fin_max},
                    })
                nueva_config["tipos_conductor"] = tipos_conductor
            
            # Consolidar nodos de todas las fuentes (Nodos, Vacíos, Desplazamientos, Paradas, Depósitos)
            nueva_config["nodos"] = sorted([n for n in nodos_recolectados if n])
            logger.info(f"Nodos consolidados desde Excel de config: {len(nueva_config['nodos'])} ({sorted(nueva_config['nodos'])})")
            
            # Asegurar que siempre haya al menos un depósito
            if not nueva_config.get("depositos"):
                deposito_principal = nueva_config.get("deposito", "Deposito Pie Andino")
                nueva_config["depositos"] = [{
                    "nombre": deposito_principal,
                    "max_buses": 200,
                    "permite_recarga": True,
                    "posiciones_recarga": 0,
                    "flota_por_tipo": nueva_config.get("flota_por_tipo", {}),
                }]
            
            # Asegurar que el depósito principal esté definido
            if not nueva_config.get("deposito") or nueva_config.get("deposito") == "":
                if nueva_config.get("depositos"):
                    nueva_config["deposito"] = nueva_config["depositos"][0]["nombre"]
                else:
                    nueva_config["deposito"] = "Deposito Pie Andino"
            # Compatibilidad: eliminar campo raíz legado, la flota se define por depósito.
            nueva_config.pop("max_buses", None)
            
            return nueva_config
            
        except Exception as e:
            logger.error(f"Error leyendo configuración desde Excel: {e}")
            traceback.print_exc()
            return None

