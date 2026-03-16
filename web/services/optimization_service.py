"""
Servicio para ejecución de optimizaciones del diagramador.
Maneja la ejecución en subprocesos separados y el monitoreo del progreso.
"""

import io
import os
import shutil
import sys
import subprocess
import threading
import time
import json
import traceback
from typing import Dict, Any, Optional
from openpyxl import load_workbook

from .config_service import ConfigService
from .excel_service import ExcelTripsService
from ..utils.logging_utils import logger

# Ruta SIEMPRE en diagramador_optimizado/resultado_diagramacion.xlsx
RESULT_FILENAME = "resultado_diagramacion.xlsx"
MOTOR_FOLDER = "diagramador_optimizado"

def _get_result_path_in_motor() -> str:
    this_file = os.path.abspath(os.path.realpath(__file__))
    services_dir = os.path.dirname(this_file)
    project_root = os.path.normpath(os.path.join(services_dir, "..", ".."))
    path = os.path.normpath(os.path.join(project_root, MOTOR_FOLDER, RESULT_FILENAME))
    # Si por cualquier motivo la ruta no incluye diagramador_optimizado, forzarla
    if MOTOR_FOLDER not in path:
        path = os.path.normpath(os.path.join(project_root, MOTOR_FOLDER, RESULT_FILENAME))
    return os.path.abspath(path)


_RESULT_PATH_IN_MOTOR = _get_result_path_in_motor()


def obtener_mensaje_error_seguro(excepcion):
    """
    Obtiene el mensaje de una excepción de forma completamente segura.
    Convierte directamente a bytes y luego a ASCII, evitando str() que puede fallar.
    
    Args:
        excepcion: La excepción de la cual obtener el mensaje
        
    Returns:
        String ASCII seguro con el mensaje de error
    """
    try:
        if hasattr(excepcion, 'args') and excepcion.args:
            arg = excepcion.args[0]
            # Intentar convertir directamente a bytes sin pasar por str()
            if isinstance(arg, str):
                # Si es string, convertir directamente a bytes y luego a ASCII
                try:
                    # Codificar a UTF-8 con reemplazo de errores
                    arg_bytes = arg.encode('utf-8', errors='replace')
                    # Decodificar a ASCII reemplazando caracteres problemáticos
                    arg_ascii = arg_bytes.decode('ascii', errors='replace')
                    return arg_ascii
                except Exception:
                    # Si falla, procesar carácter por carácter
                    resultado = []
                    for char in arg:
                        try:
                            char.encode('ascii')
                            resultado.append(char)
                        except UnicodeEncodeError:
                            resultado.append('?')
                    return ''.join(resultado)
            else:
                # Si no es string, intentar convertirlo
                try:
                    arg_str = str(arg)
                    arg_bytes = arg_str.encode('utf-8', errors='replace')
                    return arg_bytes.decode('ascii', errors='replace')
                except Exception:
                    return type(excepcion).__name__
        else:
            return type(excepcion).__name__
    except Exception:
        return "Exception"


def limpiar_mensaje_para_json(mensaje) -> str:
    """
    Limpia caracteres Unicode problemáticos de un mensaje antes de serializarlo a JSON o loguearlo.
    Esta función es crítica para evitar errores de encoding.
    Esta función NUNCA debe fallar, incluso si el mensaje contiene errores de encoding.
    
    Args:
        mensaje: Mensaje a limpiar (puede ser cualquier tipo)
        
    Returns:
        Mensaje limpio con caracteres Unicode reemplazados por ASCII
    """
    try:
        # Convertir a string de forma ULTRA-SEGURA
        # NO usar str() directamente porque puede fallar con caracteres Unicode
        mensaje_str = None
        try:
            # Si es string, usar directamente pero limpiarlo
            if isinstance(mensaje, str):
                mensaje_str = mensaje
            else:
                # Si no es string, usar obtener_mensaje_error_seguro que es más seguro
                mensaje_str = obtener_mensaje_error_seguro(mensaje)
        except Exception:
            # Si falla, intentar con repr()
            try:
                mensaje_repr = repr(mensaje)
                # Limpiar repr() removiendo comillas
                if mensaje_repr.startswith("'") and mensaje_repr.endswith("'"):
                    mensaje_str = mensaje_repr[1:-1]
                elif mensaje_repr.startswith('"') and mensaje_repr.endswith('"'):
                    mensaje_str = mensaje_repr[1:-1]
                else:
                    mensaje_str = mensaje_repr
                # Limpiar escapes
                mensaje_str = mensaje_str.replace("\\'", "'").replace('\\"', '"')
            except Exception:
                try:
                    mensaje_str = type(mensaje).__name__
                except Exception:
                    return "Mensaje no pudo ser procesado"
        
        if mensaje_str is None:
            return "Mensaje no pudo ser procesado"
        
        # Reemplazar caracteres Unicode problemáticos comunes ANTES de cualquier otra operación
        # Esto es crítico porque algunos caracteres pueden causar errores durante el encoding
        # Hacer el reemplazo de forma segura, carácter por carácter si es necesario
        reemplazos = {
            '\u2713': '[OK]',  # checkmark
            '\u2714': '[OK]',  # checkmark pesado
            '\u2264': '<=',    # menor o igual
            '\u2265': '>=',    # mayor o igual
            '\u2260': '!=',    # diferente
        }
        
        # Intentar reemplazo directo primero
        try:
            for unicode_char, replacement in reemplazos.items():
                try:
                    if unicode_char in mensaje_str:
                        mensaje_str = mensaje_str.replace(unicode_char, replacement)
                except Exception:
                    continue
        except Exception:
            # Si falla el reemplazo directo, procesar carácter por carácter
            try:
                resultado_chars = []
                for char in mensaje_str:
                    if char in reemplazos:
                        resultado_chars.append(reemplazos[char])
                    else:
                        resultado_chars.append(char)
                mensaje_str = ''.join(resultado_chars)
            except Exception:
                # Si todo falla, intentar convertir a ASCII directamente
                pass  # Si falla un reemplazo, continuar con los demás
        
        # Convertir a ASCII de forma segura usando bytes directamente
        try:
            # Codificar a UTF-8 primero (más permisivo)
            mensaje_bytes = mensaje_str.encode('utf-8', errors='replace')
            # Decodificar a ASCII reemplazando caracteres problemáticos
            mensaje_str = mensaje_bytes.decode('ascii', errors='replace')
        except Exception:
            # Si falla, procesar carácter por carácter
            try:
                resultado = []
                for char in mensaje_str:
                    try:
                        char.encode('ascii')
                        resultado.append(char)
                    except UnicodeEncodeError:
                        resultado.append('?')
                mensaje_str = ''.join(resultado)
            except Exception:
                return "Mensaje no pudo ser procesado"
        
        return mensaje_str
    except Exception:
        # Si TODO falla, devolver un mensaje genérico seguro
        return "Mensaje no pudo ser procesado"


class OptimizationService:
    """
    Servicio para ejecutar optimizaciones del diagramador.
    Usa la carpeta del motor (ej. diagramador_optimizado/) para config, Excel y resultado.
    """
    
    RESULT_FILE = "resultado_diagramacion.xlsx"
    TIMEOUT_SECONDS = 600  # 10 minutos
    
    def __init__(
        self,
        config_service: ConfigService,
        excel_trips_service: ExcelTripsService,
        motor_dir: str = "diagramador_optimizado",
        result_file_path: Optional[str] = None,
    ):
        """
        Inicializa el servicio de optimización.
        
        Args:
            config_service: Instancia de ConfigService (debe usar config_path dentro de motor_dir)
            excel_trips_service: Instancia de ExcelTripsService (debe usar trips en motor_dir)
            motor_dir: Carpeta del motor (ej. diagramador_optimizado/) donde están config, Excel y resultado
            result_file_path: Ruta absoluta del archivo de resultado. Si se indica, se usa siempre;
                si no, se usa join(motor_dir, RESULT_FILE) y se resuelve con el CWD.
        """
        self.config_service = config_service
        self.excel_trips_service = excel_trips_service
        self.motor_dir = motor_dir
        self._result_file = os.path.join(motor_dir, self.RESULT_FILE)
        # Usar result_file_path si se indica (ej. desde app.py); si no, resolver desde __file__
        if result_file_path and os.path.isabs(result_file_path):
            self._result_file_abs = os.path.normpath(os.path.abspath(result_file_path))
        else:
            self._result_file_abs = _get_result_path_in_motor()
        try:
            logger.info(f"OptimizationService: ruta resultado fijada a {self._result_file_abs}")
        except Exception:
            pass
        self._optimization_running = False
        self._optimization_thread: Optional[threading.Thread] = None
        self._last_run_status: Dict[str, Any] = {"success": None, "message": ""}
        self._lock = threading.RLock()
    
    def _get_result_path_from_config(self) -> str:
        """Ruta absoluta del resultado: SIEMPRE diagramador_optimizado/resultado_diagramacion.xlsx (desde __file__ del módulo, sin depender de config)."""
        return _get_result_path_in_motor()

    def _resolve_result_path(self, result_file_path: Optional[str] = None) -> str:
        """Ruta absoluta del resultado: SIEMPRE en diagramador_optimizado (se ignora ruta que no lo tenga)."""
        if result_file_path and os.path.isabs(result_file_path):
            norm = os.path.normpath(result_file_path)
            if self.motor_dir in norm:
                return result_file_path
        return self._get_result_path_from_config()

    def _get_result_path(self) -> str:
        """Ruta del archivo de resultado (absoluta, misma carpeta que configuracion.json)."""
        if self._result_file_abs:
            return self._result_file_abs
        return self._get_result_path_from_config()

    def get_canonical_result_path(self) -> str:
        """Ruta fija del resultado: diagramador_optimizado/resultado_diagramacion.xlsx (la de __init__)."""
        return self._result_file_abs

    def get_result_search_paths(self) -> list:
        """ÚNICA ruta donde buscar: diagramador_optimizado/resultado_diagramacion.xlsx (la de __init__, siempre correcta)."""
        return [self._result_file_abs]
    
    def start_optimization(self) -> Dict[str, Any]:
        """
        Inicia una optimización en un hilo separado.
        
        Returns:
            Diccionario con resultado de la operación
        """
        with self._lock:
            # Guardarraíl: si existe un hilo vivo, bloquear nuevos inicios aunque
            # un reset previo haya dejado _optimization_running en False.
            hilo_vivo_global = self._optimization_thread is not None and self._optimization_thread.is_alive()
            if hilo_vivo_global:
                self._optimization_running = True
                mensaje = "Ya hay una optimizacion en ejecucion"
                try:
                    mensaje = str(mensaje).encode('ascii', errors='replace').decode('ascii')
                except Exception:
                    mensaje = "Ya hay una optimizacion en ejecucion"
                return {
                    "success": False,
                    "message": mensaje
                }
            # Si el flag está True pero no hay hilo vivo (hilo terminó, crasheó o nunca se guardó), permitir ejecutar
            if self._optimization_running:
                hilo_vivo = self._optimization_thread is not None and self._optimization_thread.is_alive()
                if not hilo_vivo:
                    self._optimization_running = False
                    self._optimization_thread = None
                    logger.info("Estado 'en ejecucion' reseteado automaticamente (no hay hilo en curso).")
            if self._optimization_running:
                mensaje = "Ya hay una optimizacion en ejecucion"
                try:
                    mensaje = str(mensaje).encode('ascii', errors='replace').decode('ascii')
                except Exception:
                    mensaje = "Ya hay una optimizacion en ejecucion"
                return {
                    "success": False,
                    "message": mensaje
                }
            
            if not os.path.exists(self.config_service.config_path):
                mensaje = f"No se encontro el archivo de configuracion: {self.config_service.config_path}"
                try:
                    mensaje = str(mensaje).encode('ascii', errors='replace').decode('ascii')
                except Exception:
                    mensaje = "No se encontro el archivo de configuracion"
                return {
                    "success": False,
                    "message": mensaje
                }
            
            # Obtener ruta FRESCA del archivo Excel (sin caché) en el momento de ejecutar
            archivo_excel = self.excel_trips_service.get_trips_file_path()
            if not archivo_excel:
                mensaje = "No se encontro ningun archivo Excel cargado. Por favor, carga un archivo Excel primero."
                try:
                    mensaje = str(mensaje).encode('ascii', errors='replace').decode('ascii')
                except Exception:
                    mensaje = "No se encontro ningun archivo Excel cargado. Por favor, carga un archivo Excel primero."
                return {
                    "success": False,
                    "message": mensaje
                }
            
            # La preparación pesada (_prepare_before_optimization) se ejecuta en el hilo
            # para que esta petición responda de inmediato y la web no se quede colgada.
            
            # Iniciar optimización en hilo separado
            self._optimization_running = True
            mensaje_curso = "Optimizacion en curso"
            try:
                mensaje_curso = str(mensaje_curso).encode('ascii', errors='replace').decode('ascii')
            except Exception:
                mensaje_curso = "Optimizacion en curso"
            self._last_run_status = {"success": None, "message": mensaje_curso}
            
            thread = threading.Thread(
                target=self._execute_optimization_thread,
                args=(archivo_excel,)
            )
            self._optimization_thread = thread
            thread.start()
            
            mensaje_iniciado = "Optimizacion iniciada en un proceso separado"
            try:
                mensaje_iniciado = str(mensaje_iniciado).encode('ascii', errors='replace').decode('ascii')
            except Exception:
                mensaje_iniciado = "Optimizacion iniciada en un proceso separado"
            return {
                "success": True,
                "message": mensaje_iniciado
            }
    
    def _prepare_before_optimization(self, archivo_excel: str) -> None:
        """
        Prepara el entorno antes de ejecutar la optimización.
        
        Args:
            archivo_excel: Ruta al archivo Excel de viajes
            
        Raises:
            ValueError: Si no se pueden detectar nodos del archivo Excel
        """
        try:
            # Cargar configuración
            self.config_service.load_config()
            
            # Procesar Excel para actualizar nodos dinámicamente
            nodos_detectados = self.excel_trips_service._detect_nodes_from_excel(archivo_excel)
            if not nodos_detectados:
                error_msg = (
                    "No se pudieron detectar nodos en el archivo Excel. "
                    "Asegúrese de que el archivo tenga columnas 'Origen' y 'Destino' con datos válidos."
                )
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            # Actualizar nodos en la configuración
            self.config_service.update_nodes(nodos_detectados)
            self.config_service.save_config()
            logger.info(f"Nodos actualizados antes de optimizar: {len(nodos_detectados)} nodos - {nodos_detectados}")
            
        except ValueError:
            # Re-lanzar ValueError para que se maneje arriba
            raise
        except Exception as e:
            error_msg = f"Error preparando antes de optimizar: {e}"
            logger.error(error_msg)
            raise ValueError(error_msg) from e
    
    def reset_optimization_state(self) -> None:
        """
        Resetea el estado "optimización en ejecución".
        Útil cuando el hilo terminó con error o quedó bloqueado y la interfaz sigue mostrando "en ejecución".
        """
        with self._lock:
            # Nunca resetear mientras el hilo real sigue vivo: evita ejecuciones solapadas.
            if self._optimization_thread is not None and self._optimization_thread.is_alive():
                logger.warning("Reset de optimizacion ignorado: hay un hilo de optimizacion en ejecucion.")
                return
            self._optimization_running = False
            self._optimization_thread = None

    def _execute_optimization_thread(self, archivo_excel: str) -> None:
        """
        Ejecuta la optimización en un hilo separado.
        
        Args:
            archivo_excel: Ruta al archivo Excel de viajes
        """
        try:
            self._execute_optimization_thread_impl(archivo_excel)
        finally:
            with self._lock:
                self._optimization_running = False
                self._optimization_thread = None

    def _execute_optimization_thread_impl(self, archivo_excel: str) -> None:
        """Implementación real del hilo de optimización. Siempre se llama desde _execute_optimization_thread con finally que resetea el flag."""
        # Guardar el directorio actual ANTES de cualquier operación
        cwd_actual = os.getcwd()
        
        # Variable para almacenar el resultado del subprocess
        result = None
        
        try:
            # Preparación pesada (lee Excel, actualiza nodos) - evita bloquear la petición POST
            try:
                self._prepare_before_optimization(archivo_excel)
            except ValueError as e:
                mensaje_error = str(e)
                try:
                    mensaje_error = mensaje_error.encode('ascii', errors='replace').decode('ascii')
                except Exception:
                    mensaje_error = "Error al preparar la optimizacion"
                logger.error(mensaje_error)
                self._set_status(False, mensaje_error)
                return
            except Exception as e_prep:
                mensaje_error = str(e_prep) if str(e_prep) else "Error inesperado al preparar"
                try:
                    mensaje_error = mensaje_error.encode('ascii', errors='replace').decode('ascii')
                except Exception:
                    mensaje_error = "Error al preparar la optimizacion"
                logger.error(mensaje_error)
                self._set_status(False, mensaje_error)
                return
            
            try:
                logger.info("=" * 80)
                logger.info("PREPARANDO SUBPROCESO (ENTORNO LIMPIO)...")
                logger.info("=" * 80)
            except Exception:
                # Si falla el logueo inicial, continuar de todas formas
                pass
            
            # Verificar archivo Excel
            if not os.path.exists(archivo_excel):
                mensaje_error = f"ERROR: El archivo Excel no existe: {archivo_excel}"
                logger.error(mensaje_error)
                self._set_status(False, mensaje_error)
                return
            
            fecha_modificacion = os.path.getmtime(archivo_excel)
            fecha_modificacion_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(fecha_modificacion))
            logger.info(f"Archivo Excel encontrado: {archivo_excel}")
            logger.info(f"  - Última modificación: {fecha_modificacion_str}")
            
            # Verificar configuración
            config_path = os.path.abspath(self.config_service.config_path)
            if not os.path.exists(config_path):
                mensaje_error = f"ERROR CRÍTICO: El archivo de configuración no existe: {config_path}"
                logger.error(mensaje_error)
                self._set_status(False, mensaje_error)
                return
            
            # Leer configuración para log
            try:
                with open(config_path, "r", encoding="utf-8") as f:
                    config_log = json.load(f)
                logger.info(f"CONFIGURACIÓN ANTES DE EJECUTAR:")
                logger.info(f"  - Depósito: {config_log.get('deposito', 'N/A')}")
                logger.info(f"  - Límite jornada: {config_log.get('limite_jornada', 'N/A')} min")
                logger.info(f"  - Nodos: {config_log.get('nodos', [])} (total: {len(config_log.get('nodos', []))})")
                logger.info(f"  - Archivo Excel: {archivo_excel}")
            except Exception as e_log:
                logger.warning(f"Error leyendo config para log: {e_log}")
            
            # Ruta del resultado: la de __init__ (siempre diagramador_optimizado/resultado_diagramacion.xlsx)
            resultado_path = self._result_file_abs
            logger.info(f"RUTA_RESULTADO (diagramador_optimizado): {resultado_path}")
            if os.path.exists(resultado_path):
                try:
                    os.remove(resultado_path)
                    logger.info("Resultado anterior eliminado antes de ejecutar.")
                except Exception as e_remove:
                    logger.warning(f"No se pudo eliminar resultado anterior: {e_remove}")
            
            # Pausa para asegurar escritura completa
            time.sleep(1.0)
            
            # Usar SIEMPRE ruta absoluta para evitar leer archivo equivocado por cambios de directorio
            archivo_absoluto = os.path.abspath(archivo_excel)
            cwd_actual = os.getcwd()

            # Verificar que el archivo contiene los datos recién cargados (lectura fresca desde disco)
            try:
                import pandas as pd
                df_verif = pd.read_excel(archivo_absoluto)
                n_filas = len(df_verif) if not df_verif.empty else 0
                logger.info(f"Archivo Excel verificado: {n_filas} filas de viajes (lectura fresca desde disco)")
            except Exception as e_verif:
                logger.warning(f"No se pudo verificar conteo de viajes en Excel: {e_verif}")
            
            # Verificar si estamos en un ejecutable compilado
            es_ejecutable = getattr(sys, 'frozen', False)
            config_path_abs = os.path.abspath(self.config_service.config_path)

            if es_ejecutable:
                directorio_ejecutable = os.path.dirname(sys.executable)
                logger.info(f"Modo ejecutable detectado. Directorio del .exe: {directorio_ejecutable}")
                directorio_base = directorio_ejecutable
            else:
                # Raíz del proyecto desde este archivo (web/services/optimization_service.py -> .. -> .. = raíz)
                # Así se usa SIEMPRE el diagramador_optimizado del mismo proyecto que la web
                _this_file = os.path.abspath(os.path.realpath(__file__))
                _raiz_desde_web = os.path.normpath(os.path.join(os.path.dirname(_this_file), "..", ".."))
                if os.path.isdir(os.path.join(_raiz_desde_web, self.motor_dir)):
                    directorio_base = _raiz_desde_web
                else:
                    directorio_base = os.path.dirname(os.path.dirname(config_path_abs))
                if not os.path.isdir(os.path.join(directorio_base, self.motor_dir)):
                    directorio_base = cwd_actual
                logger.info(f"Modo desarrollo: directorio base = {directorio_base} (raíz del proyecto)")
            
            # Ejecutar SIEMPRE en subproceso aislado para evitar estado compartido
            # entre corridas web (caches/módulos en memoria) y mantener paridad con CLI.
            ejecutar_en_proceso = False
            if ejecutar_en_proceso:
                logger.info("=" * 80)
                logger.info("EJECUTANDO DIAGRAMADOR EN PROCESO (web -> diagramador_optimizado)")
                logger.info("=" * 80)
                # Usar EXACTAMENTE los mismos archivos que el CLI (mismo directorio motor)
                # para que el resultado sea idéntico a ejecutar: python -m diagramador_optimizado.cli.main
                directorio_motor = os.path.join(directorio_base, self.motor_dir)
                archivo_excel_param = os.path.abspath(os.path.join(directorio_motor, "datos_salidas.xlsx"))
                archivo_config_param = os.path.abspath(os.path.join(directorio_motor, "configuracion.json"))
                archivo_salida_param = os.path.abspath(os.path.join(directorio_motor, "resultado_diagramacion.xlsx"))
                logger.info(f"  Archivo Excel: {archivo_excel_param}")
                logger.info(f"  Configuración: {archivo_config_param}")
                logger.info(f"  Directorio base: {directorio_base}")
                logger.info(f"  Motor: diagramador_optimizado.cli.main (mismo que python -m diagramador_optimizado.cli.main)")
                logger.info("=" * 80)
                
                try:
                    # No cambiar cwd: el CLI tampoco lo hace; así el resultado es idéntico a ejecución por consola.
                    if directorio_base not in sys.path:
                        sys.path.insert(0, directorio_base)
                        logger.info(f"Raíz del proyecto añadida a sys.path: {directorio_base}")
                    
                    # Punto de entrada del optimizador: mismo que CLI (diagramador_optimizado.cli.main)
                    from diagramador_optimizado.cli.main import main
                    
                    if not os.path.exists(archivo_excel_param):
                        mensaje_error = f"ERROR: El archivo Excel no existe: {archivo_excel_param}"
                        logger.error(mensaje_error)
                        self._set_status(False, mensaje_error)
                        return
                    if not os.path.exists(archivo_config_param):
                        mensaje_error = f"ERROR: El archivo de configuración no existe: {archivo_config_param}"
                        logger.error(mensaje_error)
                        self._set_status(False, mensaje_error)
                        return
                    
                    logger.info(f"  Excel: {archivo_excel_param}")
                    logger.info(f"  Config: {archivo_config_param}")
                    logger.info(f"  Salida: {archivo_salida_param}")
                    
                    # Asegurar que el directorio de salida existe antes de ejecutar
                    dir_salida = os.path.dirname(archivo_salida_param)
                    if dir_salida and not os.path.exists(dir_salida):
                        try:
                            os.makedirs(dir_salida, exist_ok=True)
                            logger.info(f"Directorio de salida creado: {dir_salida}")
                        except Exception as e_dir:
                            logger.warning(f"No se pudo crear directorio de salida: {e_dir}")
                    
                    # Redirigir stdout/stderr para evitar OSError [Errno 22] en Windows
                    # y para capturar la salida del diagramador (diagnóstico si no se genera archivo)
                    _stdout_orig = sys.stdout
                    _stderr_orig = sys.stderr
                    _stdout_buf = io.StringIO()
                    _stderr_buf = io.StringIO()
                    sys.stdout = _stdout_buf
                    sys.stderr = _stderr_buf
                    try:
                        main(
                            archivo_excel=archivo_excel_param,
                            archivo_config=archivo_config_param,
                            archivo_salida=archivo_salida_param,
                            random_seed=42
                        )
                    finally:
                        _salida_diag = _stdout_buf.getvalue()
                        _errores_diag = _stderr_buf.getvalue()
                        sys.stdout = _stdout_orig
                        sys.stderr = _stderr_orig
                    
                    logger.info("[OK] Diagramador optimizado completado (revisando archivo...)")
                    
                    # Esperar un momento para que el archivo se escriba completamente
                    time.sleep(2.0)
                    
                    rutas_posibles = [archivo_salida_param, self._result_file_abs]
                    resultado_encontrado = None
                    logger.info("Buscando archivo de resultado...")
                    for ruta_posible in rutas_posibles:
                        ruta_abs = os.path.abspath(ruta_posible)
                        logger.info(f"  Verificando: {ruta_abs}")
                        if os.path.exists(ruta_abs):
                            tamaño = os.path.getsize(ruta_abs)
                            if tamaño > 0:
                                resultado_encontrado = ruta_abs
                                logger.info(f"[OK] Archivo de resultado encontrado: {resultado_encontrado}")
                                logger.info(f"  Tamaño: {tamaño:,} bytes ({tamaño / 1024:.2f} KB)")
                                break
                            else:
                                logger.warning(f"  Archivo existe pero está vacío: {ruta_abs}")
                    
                    # Si no está en la ruta canónica, buscar en cwd y en diagramador_optimizado y copiar ahí
                    if resultado_encontrado is None:
                        nombre_archivo = "resultado_diagramacion.xlsx"
                        dir_canonico = os.path.dirname(self._result_file_abs)
                        carpetas_buscar = [
                            os.getcwd(),
                            dir_canonico,
                            os.path.join(os.getcwd(), self.motor_dir),
                        ]
                        for carpeta in carpetas_buscar:
                            if not carpeta:
                                continue
                            candidato = os.path.join(carpeta, nombre_archivo)
                            try:
                                if os.path.isfile(candidato) and os.path.getsize(candidato) > 0:
                                    if os.path.normpath(candidato) != os.path.normpath(self._result_file_abs):
                                        if not os.path.isdir(dir_canonico):
                                            os.makedirs(dir_canonico, exist_ok=True)
                                        shutil.copy2(candidato, self._result_file_abs)
                                        logger.info(f"[OK] Archivo encontrado en {candidato} y copiado a {self._result_file_abs}")
                                        resultado_encontrado = self._result_file_abs
                                    else:
                                        resultado_encontrado = self._result_file_abs
                                    break
                            except Exception as e_copia:
                                logger.warning(f"No se pudo copiar desde {candidato}: {e_copia}")
                            if resultado_encontrado:
                                break
                    
                    if resultado_encontrado:
                        # Verificar que sea el archivo esperado (no corrupto)
                        try:
                            from openpyxl import load_workbook
                            wb_test = load_workbook(resultado_encontrado, read_only=True)
                            hojas = wb_test.sheetnames
                            wb_test.close()
                            logger.info(f"[OK] Archivo validado correctamente. Hojas: {hojas}")
                            self._set_status(True, "Optimización completada exitosamente")
                        except Exception as e_valid:
                            logger.error(f"Archivo encontrado pero corrupto: {e_valid}")
                            self._set_status(False, f"El archivo de resultados está corrupto: {e_valid}")
                    else:
                        ruta_esperada = self._result_file_abs
                        logger.error(f"Archivo de resultado no encontrado. Ubicacion esperada: {ruta_esperada}")
                        logger.error(f"  Directorio de trabajo actual: {os.getcwd()}")
                        # Loguear salida del diagramador para diagnóstico
                        if _salida_diag:
                            try:
                                for line in _salida_diag.splitlines():
                                    logger.info(f"  [diag] {line}")
                            except Exception:
                                logger.info(f"  [diag] (salida no legible)")
                        if _errores_diag:
                            try:
                                for line in _errores_diag.splitlines():
                                    logger.error(f"  [diag stderr] {line}")
                            except Exception:
                                pass
                        # Mensaje breve para el usuario: extraer primera línea con ERROR o "No se pudo"
                        causa_breve = ""
                        for line in (_salida_diag + "\n" + _errores_diag).splitlines():
                            line = (line or "").strip()
                            if "ERROR" in line or "No se pudo" in line or "Error" in line:
                                causa_breve = line[:200]
                                break
                        if causa_breve:
                            try:
                                causa_breve = limpiar_mensaje_para_json(causa_breve)
                            except Exception:
                                pass
                        msg_usuario = (
                            causa_breve
                            if causa_breve
                            else "La optimización terminó pero no se generó el archivo."
                        )
                        if len(msg_usuario) > 350:
                            msg_usuario = msg_usuario[:347] + "..."
                        self._set_status(False, msg_usuario)
                    
                    return
                        
                except Exception as e_directo:
                    try:
                        _out = _stdout_buf.getvalue()
                        _err = _stderr_buf.getvalue()
                    except Exception:
                        _out = _err = ""
                    if _out or _err:
                        logger.info("Salida del diagramador antes del fallo:")
                        for line in (_out + "\n" + _err).splitlines():
                            if line.strip():
                                logger.info(f"  [diag] {line[:500]}")
                    error_msg = f"Error ejecutando diagramador directamente: {e_directo}"
                    logger.error(error_msg)
                    import traceback
                    logger.error(traceback.format_exc())
                    msg_limpio = limpiar_mensaje_para_json(str(e_directo))
                    if len(msg_limpio) > 400:
                        msg_limpio = msg_limpio[:397] + "..."
                    self._set_status(False, "Error en optimizacion: " + msg_limpio)
                    return
            
            # Si NO estamos en un ejecutable, usar subproceso como antes
            python_executable = sys.executable
            
            # Crear un script temporal que ejecute el diagramador completo
            # Esto asegura que se ejecute todo el diagramador optimizado (Fase 1, 2 y 3)
            import tempfile
            script_temp = tempfile.NamedTemporaryFile(
                mode='w',
                suffix='.py',
                delete=False,
                encoding='utf-8'
            )
            
            # Rutas: Excel relativo o absoluto; config y salida resueltas desde disco (misma carpeta que config)
            archivo_excel_relativo = os.path.relpath(archivo_absoluto, cwd_actual) if not os.path.isabs(archivo_absoluto) else archivo_absoluto
            archivo_config_abs = os.path.abspath(self.config_service.config_path)
            archivo_salida_abs = self._result_file_abs
            
            script_content = f"""# Script temporal para ejecutar diagramador optimizado completo
# -*- coding: utf-8 -*-
import sys
import os
import io

# CONFIGURAR CODIFICACIÓN UTF-8 LO MÁS TEMPRANO POSIBLE (CRÍTICO)
# Esto debe ser lo primero que se ejecute para evitar errores de cp1252
def _setup_encoding():
    \"\"\"Configura la codificación UTF-8 para stdout/stderr.\"\"\"
    if sys.platform != 'win32':
        return
    
    try:
        # Método 1: reconfigure (Python 3.7+)
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        if hasattr(sys.stderr, 'reconfigure'):
            sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass
    
    try:
        # Método 2: TextIOWrapper si tiene buffer
        if hasattr(sys.stdout, 'buffer'):
            sys.stdout = io.TextIOWrapper(
                sys.stdout.buffer,
                encoding='utf-8',
                errors='replace',
                line_buffering=True
            )
        if hasattr(sys.stderr, 'buffer'):
            sys.stderr = io.TextIOWrapper(
                sys.stderr.buffer,
                encoding='utf-8',
                errors='replace',
                line_buffering=True
            )
    except Exception:
        pass

# Ejecutar configuración de codificación inmediatamente
_setup_encoding()

# Cambiar al directorio de trabajo del proyecto
cwd = {repr(cwd_actual)}
os.chdir(cwd)

# Agregar directorio actual al path
sys.path.insert(0, cwd)

# Importar y ejecutar el diagramador (mismo punto de entrada que CLI)
from diagramador_optimizado.cli.main import main

# Ejecutar con parámetros (config y salida absolutas para escribir en diagramador_optimizado/)
try:
    main(
        archivo_excel={repr(archivo_excel_relativo)},
        archivo_config={repr(archivo_config_abs)},
        archivo_salida={repr(archivo_salida_abs)},
        random_seed=42
    )
    print("\\n[OK] Diagramador optimizado completado exitosamente")
except Exception as e:
    print(f"\\n[ERROR] ERROR en diagramador optimizado: {{e}}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
"""
            script_temp.write(script_content)
            script_temp.close()
            script_path = script_temp.name
            
            command_list = [
                python_executable,
                script_path,
            ]
            
            logger.info(f"Iniciando subprocess para ejecutar DIAGRAMADOR OPTIMIZADO COMPLETO:")
            logger.info(f"  Python: {python_executable}")
            logger.info(f"  Archivo Excel: {archivo_absoluto}")
            logger.info(f"  Configuración: {config_path}")
            logger.info(f"  Ejecutando: diagramador_optimizado.main.main()")
            logger.info(f"  - Fase 1 (Buses): SÍ")
            logger.info(f"  - Fase 2 (Conductores): SÍ")
            logger.info(f"  - Fase 3 (Unión Conductores): Según configuración")
            logger.info("=" * 80)
            
            # Ejecutar subproceso
            logger.info(f"Ejecutando diagramador con timeout de {self.TIMEOUT_SECONDS} segundos...")
            
            try:
                
                # Configurar entorno para usar UTF-8 (evitar problemas de codificación en Windows)
                env = os.environ.copy()
                env['PYTHONIOENCODING'] = 'utf-8'
                env['PYTHONUTF8'] = '1'
                env['PYTHONLEGACYWINDOWSSTDIO'] = '0'
                # Forzar UTF-8 con reemplazo de caracteres problemáticos
                if sys.platform == 'win32':
                    env['PYTHONIOENCODING'] = 'utf-8:replace'
                    # También configurar código de página de consola en Windows
                    try:
                        import subprocess as sp
                        # Establecer código de página UTF-8 en la consola
                        sp.run(['chcp', '65001'], shell=True, capture_output=True, check=False)
                    except Exception:
                        pass
                
                # Ejecutar subproceso con Popen para streaming de logs
                process = subprocess.Popen(
                    command_list,
                    cwd=cwd_actual,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,  # Redirigir stderr a stdout para leer ambos juntos
                    text=True,
                    encoding='utf-8',
                    errors='replace',
                    env=env,
                    bufsize=1,  # Line buffered
                    universal_newlines=True
                )
                
                logger.info("--- Iniciando captura de salida en tiempo real ---")
                
                # Leer salida en tiempo real
                salida_stream: list[str] = []
                while True:
                    line = process.stdout.readline()
                    if not line and process.poll() is not None:
                        break
                    if line:
                        line_text = line.strip()
                        if line_text:
                            salida_stream.append(line_text)
                            # Limpiar caracteres Unicode antes de loguear
                            try:
                                line_clean = line_text.encode('ascii', errors='replace').decode('ascii')
                                logger.info(line_clean)
                            except Exception:
                                pass
                
                # Esperar a que termine con timeout
                try:
                    stdout, stderr = process.communicate(timeout=self.TIMEOUT_SECONDS)
                    # En este punto, stdout/stderr ya están capturados por el bucle anterior
                    # ya que redirigimos stderr a stdout.
                except subprocess.TimeoutExpired:
                    process.kill()
                    logger.error("=" * 80)
                    logger.error("ERROR: El diagramador excedió el tiempo límite (600s)")
                    self._set_status(False, "La optimización excedió el tiempo límite.")
                    return

                # Crear un objeto similar al resultado de subprocess.run para compatibilidad
                class ProcessResult:
                    def __init__(self, returncode, stdout, stderr):
                        self.returncode = returncode
                        self.stdout = stdout
                        self.stderr = stderr
                
                result = ProcessResult(process.returncode, "", "")
            except subprocess.TimeoutExpired:
                logger.error("=" * 80)
                logger.error("ERROR: El diagramador excedió el tiempo límite")
                logger.error(f"El proceso se canceló después de {self.TIMEOUT_SECONDS} segundos")
                self._set_status(
                    False,
                    "La optimización excedió el tiempo límite (10 min). Ajusta la configuración o reduce datos."
                )
                return
            except Exception as e_subprocess:
                logger.error(f"ERROR al ejecutar subproceso: {e_subprocess}")
                self._set_status(False, f"Error al ejecutar el solver: {e_subprocess}")
                return
            
            # ========================================================================
            # CRITICO: Verificar el archivo INMEDIATAMENTE - PRIMERO, ANTES DE TODO
            # ========================================================================
            # Esto debe ser lo PRIMERO que hacemos después de que termine el subprocess
            # NO usar logger aquí para evitar cualquier error de Unicode
            # Si el archivo existe, marcamos como completado y retornamos INMEDIATAMENTE
            rutas_criticas = self.get_result_search_paths()
            # Buscar el archivo SIN usar logger (para evitar errores de Unicode)
            archivo_encontrado = False
            for ruta_verificar in rutas_criticas:
                try:
                    if os.path.exists(ruta_verificar):
                        size = os.path.getsize(ruta_verificar)
                        if size > 0:
                            # ARCHIVO ENCONTRADO - marcar como completado INMEDIATAMENTE
                            # NO usar logger aquí para evitar errores de Unicode
                            self._set_status(True, "Optimizacion completada correctamente.")
                            archivo_encontrado = True
                            break
                except Exception:
                    continue  # Continuar buscando en otras rutas
            
            # Si encontramos el archivo, retornar INMEDIATAMENTE sin procesar nada más
            if archivo_encontrado:
                return
            
            # Loguear salida (completamente opcional - si falla, simplemente continuar)
            # El procesamiento de stdout es solo informativo, no crítico
            # Si falla, no debe interrumpir la verificación del archivo
            try:
                try:
                    logger.info("=" * 80)
                except:
                    pass
                try:
                    logger.info("--- Salida STDOUT del diagramador ---")
                except:
                    pass
                
                if result.stdout:
                    # Procesar stdout línea por línea de forma completamente segura
                    # PRIMERO: Decodificar y limpiar el stdout completo ANTES de dividirlo
                    try:
                        # Decodificar stdout de forma segura
                        if isinstance(result.stdout, bytes):
                            stdout_text = result.stdout.decode('utf-8', errors='replace')
                        else:
                            stdout_text = str(result.stdout)
                    except Exception:
                        try:
                            if isinstance(result.stdout, bytes):
                                stdout_text = result.stdout.decode('latin-1', errors='replace')
                            else:
                                stdout_text = str(result.stdout)
                        except Exception:
                            stdout_text = "[stdout no pudo ser decodificado]"
                    
                    # Limpiar el texto completo ANTES de dividirlo en líneas
                    try:
                        # Reemplazar caracteres Unicode problemáticos comunes ANTES de cualquier otra operación
                        stdout_text = stdout_text.replace('\u2264', '<=').replace('\u2265', '>=').replace('\u2260', '!=')
                        stdout_text = stdout_text.replace('\u2713', '[OK]').replace('\u2714', '[OK]')
                        # Convertir a ASCII de forma agresiva
                        stdout_text = stdout_text.encode('ascii', errors='replace').decode('ascii')
                    except Exception:
                        # Si falla, intentar limpiar carácter por carácter
                        try:
                            resultado_chars = []
                            for char in stdout_text:
                                if char in ['\u2264', '\u2265', '\u2260', '\u2713', '\u2714']:
                                    if char == '\u2264':
                                        resultado_chars.append('<=')
                                    elif char == '\u2265':
                                        resultado_chars.append('>=')
                                    elif char == '\u2260':
                                        resultado_chars.append('!=')
                                    else:
                                        resultado_chars.append('[OK]')
                                else:
                                    try:
                                        char.encode('ascii')
                                        resultado_chars.append(char)
                                    except UnicodeEncodeError:
                                        resultado_chars.append('?')
                            stdout_text = ''.join(resultado_chars)
                        except Exception:
                            stdout_text = "[stdout no pudo ser limpiado]"
                    
                    # Ahora dividir en líneas y procesar
                    lineas_procesadas = 0
                    for linea in stdout_text.split('\n'):
                        if linea.strip():
                            try:
                                # La línea ya está limpia, solo asegurarse de que sea string
                                linea_limpia = str(linea) if not isinstance(linea, str) else linea
                                # Limpiar una vez más por si acaso
                                try:
                                    linea_limpia = linea_limpia.encode('ascii', errors='replace').decode('ascii')
                                except Exception:
                                    pass
                                # Intentar loguear - si falla, simplemente continuar
                                try:
                                    logger.info(linea_limpia)
                                    lineas_procesadas += 1
                                except:
                                    # Si falla el logueo, simplemente continuar con la siguiente línea
                                    pass
                            except Exception:
                                # Si falla cualquier cosa, simplemente continuar
                                pass
                    
                    if lineas_procesadas == 0:
                        try:
                            logger.info("[STDOUT procesado pero no se pudo loguear completamente]")
                        except:
                            pass
                else:
                    try:
                        logger.info("[Sin salida STDOUT]")
                    except:
                        pass
                
                try:
                    logger.info("--- Fin STDOUT ---")
                except:
                    pass
                try:
                    logger.info("=" * 80)
                except:
                    pass
            except Exception:
                # Si falla completamente el procesamiento de stdout, simplemente continuar
                # No es crítico, solo informativo
                pass
            
            if result.returncode != 0:
                logger.error(f"ERROR: El diagramador falló (código: {result.returncode})")
                logger.error("--- Salida STDERR ---")
                if result.stderr:
                    for linea in result.stderr.split('\n'):
                        if linea.strip():
                            logger.error(linea)
                else:
                    logger.error("[Sin salida STDERR]")
                logger.error("--- Fin STDERR ---")
                
                # Extraer mensaje de error más descriptivo
                if result.stderr and result.stderr.strip():
                    # Buscar líneas clave del error
                    lineas_error = result.stderr.strip().splitlines()
                    error_principal = None
                    
                    # Buscar errores de sintaxis primero
                    for linea in reversed(lineas_error):
                        if "IndentationError" in linea or "SyntaxError" in linea:
                            error_principal = f"Error de sintaxis en el código: {linea}"
                            break
                        elif "File" in linea and "line" in linea:
                            error_principal = f"Error en archivo: {linea}"
                            break
                        elif "Error" in linea or "ERROR" in linea:
                            if not error_principal:
                                error_principal = linea
                    
                    # Si no encontramos un error específico, usar el último
                    if not error_principal:
                        error_principal = lineas_error[-1] if lineas_error else "Error desconocido"
                else:
                    # stderr puede venir vacío porque se redirige a stdout en streaming.
                    error_principal = None
                    try:
                        for linea in reversed(salida_stream):
                            l = (linea or "").strip()
                            if not l:
                                continue
                            if (
                                "Traceback" in l
                                or "ValueError" in l
                                or "RuntimeError" in l
                                or "ERROR" in l
                                or "Error" in l
                            ):
                                error_principal = l
                                break
                        if not error_principal and salida_stream:
                            error_principal = salida_stream[-1]
                    except Exception:
                        error_principal = None
                    if not error_principal:
                        error_principal = "Error desconocido (sin detalles en STDERR/STDOUT)"
                
                mensaje_error = f"La optimización falló antes de generar resultados: {error_principal}"
                logger.error(f"Mensaje de error final para el usuario: {mensaje_error}")
                self._set_status(False, mensaje_error)
            else:
                logger.info("[OK] Subprocess completado exitosamente.")
                resultado_path = self._result_file_abs
                logger.info(f"Verificando archivo de resultado en: {resultado_path}")
                logger.info(f"Directorio de trabajo actual: {os.getcwd()}")
                
                if os.path.exists(resultado_path):
                    size = os.path.getsize(resultado_path)
                    logger.info(f"[OK] Archivo de resultado generado: {size} bytes")
                    self._set_status(True, "Optimización completada correctamente.")
                else:
                    rutas_posibles = [self._result_file_abs]
                    archivo_encontrado = False
                    for ruta in rutas_posibles:
                        if os.path.exists(ruta):
                            logger.info(f"[OK] Archivo encontrado en ubicacion alternativa: {ruta}")
                            # Copiar o mover a la ubicación esperada si es necesario
                            try:
                                import shutil
                                if ruta != resultado_path:
                                    shutil.copy2(ruta, resultado_path)
                                    logger.info(f"Archivo copiado a la ubicación esperada: {resultado_path}")
                                size = os.path.getsize(resultado_path)
                                logger.info(f"[OK] Archivo de resultado generado: {size} bytes")
                                self._set_status(True, "Optimización completada correctamente.")
                                archivo_encontrado = True
                                break
                            except Exception as e_copy:
                                logger.warning(f"No se pudo copiar archivo: {e_copy}")
                    
                    if not archivo_encontrado:
                        ruta_esperada = self._result_file_abs
                        advertencia = (
                            f"Subprocess termino bien pero no se encontro resultado. "
                            f"Ubicacion esperada: {ruta_esperada}. "
                            f"Cierra el archivo si estaba abierto y usa el boton Descargar Resultado Excel."
                        )
                        logger.warning(f"ADVERTENCIA: {advertencia}")
                        self._set_status(False, advertencia)
                    
        except Exception as e:
            # CRITICO: ANTES de hacer CUALQUIER otra cosa, verificar si el archivo se generó
            # Esto es lo MÁS IMPORTANTE - el archivo puede existir aunque haya errores en los logs
            archivo_generado = False
            
            # Verificación ULTRA-SEGURA sin usar logger (para evitar errores de Unicode)
            try:
                resultado_path = self._get_result_path()
                rutas_verificar = self.get_result_search_paths()
                for ruta in rutas_verificar:
                    try:
                        if os.path.exists(ruta):
                            size = os.path.getsize(ruta)
                            if size > 0:  # Verificar que el archivo no esté vacío
                                # Archivo encontrado - marcar como completado INMEDIATAMENTE
                                # NO usar logger aquí para evitar cualquier error de Unicode
                                self._set_status(True, "Optimizacion completada correctamente.")
                                archivo_generado = True
                                
                                # Intentar copiar a la ubicación esperada si es necesario
                                try:
                                    import shutil
                                    if ruta != resultado_path:
                                        shutil.copy2(ruta, resultado_path)
                                except Exception:
                                    pass  # Si no se puede copiar, no importa
                                
                                break  # Archivo encontrado, salir del loop
                    except Exception:
                        continue  # Continuar buscando en otras rutas
            except Exception:
                # Si falla completamente la verificación, continuar con manejo de errores
                pass
            
            # Si encontramos el archivo, retornar inmediatamente sin procesar el error
            if archivo_generado:
                return
            
            # Solo si NO encontramos el archivo, procesar el error
            # Importar traceback explícitamente aquí para evitar problemas de scope
            import traceback as tb_module
            
            try:
                try:
                    # Obtener mensaje de error de forma segura usando la función helper
                    error_msg = obtener_mensaje_error_seguro(e)
                    # Limpiar adicionalmente - hacerlo múltiples veces para asegurar
                    error_msg = limpiar_mensaje_para_json(error_msg)
                    # Limpiar una vez más para asegurar que no queden caracteres problemáticos
                    try:
                        error_msg = error_msg.encode('utf-8', errors='replace').decode('ascii', errors='replace')
                    except Exception:
                        error_msg = "Error al procesar mensaje de excepcion"
                    
                    # Construir el mensaje completo de forma segura
                    try:
                        # Asegurarse de que error_msg esté completamente limpio antes de concatenar
                        error_msg_final = limpiar_mensaje_para_json(error_msg)
                        try:
                            error_msg_final = error_msg_final.encode('ascii', errors='replace').decode('ascii')
                        except Exception:
                            error_msg_final = "Error al procesar mensaje de excepcion"
                        
                        # Construir mensaje de forma segura
                        mensaje_completo = "ERROR CRITICO en el hilo de optimizacion: " + error_msg_final
                        # Limpiar el mensaje completo también
                        mensaje_completo = limpiar_mensaje_para_json(mensaje_completo)
                        try:
                            mensaje_completo = mensaje_completo.encode('ascii', errors='replace').decode('ascii')
                        except Exception:
                            mensaje_completo = "ERROR CRITICO en el hilo de optimizacion"
                        
                        logger.error(mensaje_completo)
                    except Exception:
                        # Si falla, usar un mensaje simple
                        try:
                            logger.error("ERROR CRITICO en el hilo de optimizacion")
                        except Exception:
                            # Si incluso esto falla, no hacer nada
                            pass
                    traceback_str = tb_module.format_exc()
                    # Limpiar caracteres Unicode problemáticos en el traceback
                    try:
                        traceback_limpio = traceback_str.encode('utf-8', errors='replace').decode('ascii', errors='replace')
                        logger.error("Traceback: " + traceback_limpio)
                    except Exception:
                        # Si falla, intentar limpiar carácter por carácter
                        traceback_limpio = limpiar_mensaje_para_json(traceback_str)
                        logger.error("Traceback: " + traceback_limpio)
                except Exception as e_log_error:
                    # Si incluso el logueo del error falla, intentar un mensaje simple
                    try:
                        error_simple = obtener_mensaje_error_seguro(e)
                        error_simple = limpiar_mensaje_para_json(error_simple)
                        logger.error("ERROR CRITICO: " + error_simple)
                        error_log_str = obtener_mensaje_error_seguro(e_log_error)
                        logger.error("Error adicional al loguear traceback: " + error_log_str)
                    except:
                        # Si todo falla, al menos intentar establecer el estado
                        pass
                error_final = obtener_mensaje_error_seguro(e)
                error_final = limpiar_mensaje_para_json(error_final)
                self._set_status(False, "Error inesperado en la optimizacion: " + error_final)
            except Exception:
                # Si todo falla, establecer un estado de error genérico
                try:
                    self._set_status(False, "Error inesperado en la optimizacion")
                except Exception:
                    pass
        finally:
            # Limpiar script temporal si existe
            try:
                if 'script_path' in locals() and script_path and os.path.exists(script_path):
                    os.remove(script_path)
                    logger.debug(f"Script temporal eliminado: {script_path}")
            except Exception as e_cleanup:
                logger.warning(f"No se pudo eliminar script temporal: {e_cleanup}")
    
    def _set_status(self, success: bool, message: str) -> None:
        """
        Establece el estado de la última ejecución.
        
        Args:
            success: Si la optimización fue exitosa
            message: Mensaje descriptivo
        """
        # Limpiar el mensaje de caracteres Unicode problemáticos ANTES de guardarlo
        # Esto previene errores cuando se serializa a JSON o se escribe en logs
        mensaje_limpio = limpiar_mensaje_para_json(message)
        
        with self._lock:
            self._last_run_status = {"success": success, "message": mensaje_limpio}
    
    def get_progress(self) -> Dict[str, Any]:
        """
        Obtiene el progreso de la optimización actual.
        
        Returns:
            Diccionario con información de progreso
        """
        with self._lock:
            if self._optimization_running:
                mensaje_ejecutando = limpiar_mensaje_para_json("Ejecutando optimizacion...")
                return {
                    "success": True,
                    "is_running": True,
                    "progress": 50,
                    "message": mensaje_ejecutando,
                }
            else:
                # Si la última ejecución falló explícitamente, devolver ese error
                # (no reportar éxito por un archivo de una ejecución anterior)
                if self._last_run_status.get("success") is False:
                    mensaje = self._last_run_status.get("message", "Optimizacion finalizada con errores.")
                    mensaje = limpiar_mensaje_para_json(mensaje)
                    return {
                        "success": False,
                        "is_running": False,
                        "progress": 0,
                        "message": mensaje,
                    }
                # Buscar archivo en múltiples ubicaciones para verificar si existe
                archivo_encontrado = False
                rutas_verificar = self.get_result_search_paths()
                for ruta in rutas_verificar:
                    try:
                        if os.path.exists(ruta) and os.path.getsize(ruta) > 0:
                            archivo_encontrado = True
                            break
                    except Exception:
                        continue
                # Si el archivo existe y no hubo fallo explícito, reportar éxito
                if archivo_encontrado:
                    # Actualizar el estado sin llamar a _set_status (evitar deadlock: ya tenemos el lock)
                    if self._last_run_status.get("success") is not True:
                        msg = limpiar_mensaje_para_json("Optimizacion completada correctamente (archivo encontrado).")
                        self._last_run_status = {"success": True, "message": msg}
                    mensaje_exito = limpiar_mensaje_para_json("Optimizacion completada correctamente. Archivo generado exitosamente.")
                    return {
                        "success": True,
                        "is_running": False,
                        "progress": 100,
                        "message": mensaje_exito,
                    }
                else:
                    # Si no hay archivo, usar el mensaje del estado
                    mensaje = self._last_run_status.get("message", "Optimizacion finalizada (sin resultados)")
                    # Si el mensaje es el antiguo (Rutas verificadas en raíz), reemplazar por mensaje con ruta correcta
                    if "Rutas verificadas" in (mensaje or "") or ("resultado_diagramacion.xlsx" in (mensaje or "") and "diagramador_optimizado" not in (mensaje or "")):
                        ruta_ok = self._result_file_abs
                        mensaje = (
                            f"Subprocess termino bien pero no se encontro resultado. "
                            f"Ubicacion esperada: {ruta_ok}. "
                            f"Cierra el archivo si estaba abierto y usa el boton Descargar Resultado Excel."
                        )
                    mensaje = limpiar_mensaje_para_json(mensaje)
                    return {
                        "success": False,
                        "is_running": False,
                        "progress": 0,
                        "message": mensaje,
                    }
    
    def get_results_summary(self) -> Dict[str, Any]:
        """
        Obtiene un resumen de los resultados de la optimización.
        
        Returns:
            Diccionario con resumen de resultados
        """
        try:
            # Buscar archivo en múltiples ubicaciones posibles
            rutas_posibles = self.get_result_search_paths()
            archivo_resultado = None
            for ruta in rutas_posibles:
                if os.path.exists(ruta):
                    archivo_resultado = os.path.abspath(ruta)
                    logger.info(f"Archivo de resultados encontrado en: {archivo_resultado}")
                    break
            
            if not archivo_resultado:
                logger.error(f"Archivo de resultados no encontrado en ninguna ubicación: {rutas_posibles}")
                return {
                    "success": False,
                    "message": "El archivo de resultados no existe",
                    "resultados": {"total_buses": 0, "total_conductors": 0},
                }
            
            # Usar el archivo encontrado en lugar de self.RESULT_FILE
            if not os.path.exists(archivo_resultado):
                return {
                    "success": False,
                    "message": "El archivo de resultados no existe",
                    "resultados": {"total_buses": 0, "total_conductors": 0},
                }
            
            wb = load_workbook(archivo_resultado, data_only=True)
            resumen = {"total_buses": 0, "total_conductors": 0}
            
            # Leer buses
            if "BloquesBuses" in wb.sheetnames:
                ws_buses = wb["BloquesBuses"]
                buses_unicos = set()
                for row in ws_buses.iter_rows(min_row=2, values_only=True):
                    if row and row[0] is not None:
                        bus_id = row[0]
                        if isinstance(bus_id, (int, float)):
                            buses_unicos.add(int(bus_id))
                        elif bus_id:
                            buses_unicos.add(bus_id)
                resumen["total_buses"] = len(buses_unicos)
            
            # Leer conductores
            if "TurnosConductores" in wb.sheetnames:
                ws_conductores = wb["TurnosConductores"]
                conductores_unicos = set()
                for row in ws_conductores.iter_rows(min_row=2, values_only=True):
                    if row and row[0] is not None:
                        conductor_id = row[0]
                        if isinstance(conductor_id, (int, float)):
                            conductores_unicos.add(int(conductor_id))
                        elif conductor_id:
                            conductores_unicos.add(conductor_id)
                resumen["total_conductors"] = len(conductores_unicos)
            
            return {"success": True, "resultados": resumen}
            
        except Exception as e:
            error_msg = str(e)
            logger.error(f"Error obteniendo resultados: {error_msg}")
            
            if "BadZipFile" in error_msg or "not a zip file" in error_msg.lower():
                return {
                    "success": False,
                    "message": "El archivo Excel generado está corrupto.",
                    "resultados": {"total_buses": 0, "total_conductors": 0},
                }
            
            return {
                "success": False,
                "message": f"Error leyendo resultados: {error_msg}",
                "resultados": {"total_buses": 0, "total_conductors": 0},
            }
    
    def get_result_file(self) -> str:
        """
        Obtiene la ruta del archivo de resultados.
        
        Returns:
            Ruta del archivo de resultados
        """
        return self._get_result_path()

