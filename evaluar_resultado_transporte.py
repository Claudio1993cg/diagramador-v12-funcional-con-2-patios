import os
from collections import defaultdict, Counter
from openpyxl import load_workbook


def to_min(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    if ":" in s:
        try:
            h, m = s.split(":")
            return int(h) * 60 + int(m)
        except Exception:
            return None
    try:
        return int(float(s))
    except Exception:
        return None


def norm_node(v):
    return (str(v or "").strip().upper())


def canon_node(v):
    # Canonicaliza alias de depósito para evitar falsos teletransportes
    # (ej: "AGUIRRE LUCO" vs "DEPOSITO AGUIRRE LUCO").
    return " ".join(norm_node(v).replace("DEPOSITO", "").split())


def analizar_secuencia(eventos, key_id, tipo_col, ini_col, fin_col, orig_col, dest_col, ignorar_tipos=None):
    ignorar_tipos = set(ignorar_tipos or [])
    por_id = defaultdict(list)
    for r in eventos:
        k = r[key_id]
        if k in (None, ""):
            continue
        por_id[k].append(r)

    resumen = {
        "total_ids": len(por_id),
        "ids_ok_tiempo": 0,
        "ids_ok_nodo": 0,
        "huecos": [],
        "solapes": [],
        "teletransportes": [],
    }

    for k, lista in por_id.items():
        lista = sorted(lista, key=lambda e: (to_min(e[ini_col]) or -1, to_min(e[fin_col]) or -1))
        ok_t = True
        ok_n = True
        for i in range(len(lista) - 1):
            a = lista[i]
            b = lista[i + 1]
            ta = (str(a[tipo_col] or "")).strip()
            tb = (str(b[tipo_col] or "")).strip()
            fa = to_min(a[fin_col])
            ib = to_min(b[ini_col])
            if fa is None or ib is None:
                continue
            diff = ib - fa
            if diff > 0:
                ok_t = False
                resumen["huecos"].append((k, diff, ta, tb, a[fin_col], b[ini_col]))
            elif diff < 0:
                ok_t = False
                resumen["solapes"].append((k, diff, ta, tb, a[fin_col], b[ini_col]))

            if ta in ignorar_tipos or tb in ignorar_tipos:
                continue
            da = canon_node(a[dest_col] or a[orig_col])
            ob = canon_node(b[orig_col] or b[dest_col])
            if da and ob and da != ob:
                ok_n = False
                resumen["teletransportes"].append((k, da, ob, ta, tb, a[fin_col], b[ini_col]))

        if ok_t:
            resumen["ids_ok_tiempo"] += 1
        if ok_n:
            resumen["ids_ok_nodo"] += 1

    return resumen


def main():
    root = os.path.dirname(__file__)
    xlsx = os.path.join(root, "diagramador_optimizado", "resultado_diagramacion.xlsx")
    wb = load_workbook(xlsx, data_only=True)

    sh_bus = wb["BusEventos"]
    sh_cond = wb["EventosCompletos"]

    bus_rows = [r for r in sh_bus.iter_rows(min_row=2, values_only=True)]
    cond_rows = [r for r in sh_cond.iter_rows(min_row=2, values_only=True)]

    # Validación adicional: InS/FnS siempre en patio (depósito)
    por_cond = defaultdict(list)
    for r in cond_rows:
        cid = r[2]
        if cid in (None, ""):
            continue
        por_cond[cid].append(r)
    ins_fns_fuera_patio = []
    for cid, lista in por_cond.items():
        for r in lista:
            tipo = (str(r[0] or "")).strip()
            if tipo not in ("InS", "FnS"):
                continue
            o = norm_node(r[6])
            d = norm_node(r[7])
            if "DEPOSITO" not in o or "DEPOSITO" not in d:
                ins_fns_fuera_patio.append((cid, tipo, r[3], r[4], r[6], r[7]))

    # BusEventos columns:
    # 0 Evento, 2 Inicio, 3 De, 4 Fin, 5 A, 8 Bus
    bus = analizar_secuencia(
        bus_rows, key_id=8, tipo_col=0, ini_col=2, fin_col=4, orig_col=3, dest_col=5
    )
    # EventosCompletos columns:
    # 0 Tipo, 1 Bus, 2 Conductor, 3 Inicio, 4 Fin, 6 Origen, 7 Destino
    # Regla dura: también se auditan transiciones con InS/FnS.
    # Antes se ignoraban y eso ocultaba teletransportes (p.ej. InS en depósito
    # seguido de Comercial iniciando en otro nodo a la misma hora).
    cond = analizar_secuencia(
        cond_rows, key_id=2, tipo_col=0, ini_col=3, fin_col=4, orig_col=6, dest_col=7,
        ignorar_tipos=set()
    )

    print("=== AUDITORIA TRANSPORTE ===")
    print(f"Buses auditados: {bus['total_ids']}")
    print(f"Buses sin huecos/solapes: {bus['ids_ok_tiempo']}/{bus['total_ids']}")
    print(f"Buses sin teletransportes: {bus['ids_ok_nodo']}/{bus['total_ids']}")
    print(f"Huecos bus: {len(bus['huecos'])}")
    print(f"Solapes bus: {len(bus['solapes'])}")
    print(f"Teletransportes bus: {len(bus['teletransportes'])}")
    if bus["huecos"]:
        c = Counter((x[2], x[3]) for x in bus["huecos"])
        print("Top huecos bus (tipoA->tipoB):", c.most_common(5))
        print("Ejemplos huecos bus:", bus["huecos"][:5])
    if bus["teletransportes"]:
        c = Counter((x[1], x[2]) for x in bus["teletransportes"])
        print("Top teletransportes bus (dest->orig):", c.most_common(5))

    print("---")
    print(f"Conductores auditados: {cond['total_ids']}")
    print(f"Conductores sin huecos/solapes: {cond['ids_ok_tiempo']}/{cond['total_ids']}")
    print(f"Conductores sin teletransportes: {cond['ids_ok_nodo']}/{cond['total_ids']}")
    print(f"Huecos conductor: {len(cond['huecos'])}")
    print(f"Solapes conductor: {len(cond['solapes'])}")
    print(f"Teletransportes conductor: {len(cond['teletransportes'])}")
    if cond["huecos"]:
        c = Counter((x[2], x[3]) for x in cond["huecos"])
        print("Top huecos conductor (tipoA->tipoB):", c.most_common(5))
        print("Ejemplos huecos conductor:", cond["huecos"][:5])
    if cond["solapes"]:
        c = Counter((x[2], x[3]) for x in cond["solapes"])
        print("Top solapes conductor (tipoA->tipoB):", c.most_common(5))
        print("Ejemplos solapes conductor:", cond["solapes"][:5])
    if cond["teletransportes"]:
        c = Counter((x[1], x[2]) for x in cond["teletransportes"])
        print("Top teletransportes conductor (dest->orig):", c.most_common(5))
        print("Ejemplos teletransportes conductor:", cond["teletransportes"][:5])
    print(f"InS/FnS fuera de patio: {len(ins_fns_fuera_patio)}")
    if ins_fns_fuera_patio:
        print("Ejemplos InS/FnS fuera de patio:", ins_fns_fuera_patio[:5])


if __name__ == "__main__":
    main()

